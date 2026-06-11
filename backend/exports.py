import io
import datetime
from flask import send_file
from db import get_db, dicts

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL = True
except ImportError:
    OPENPYXL = False

# ── Colors ──
GREEN_FILL = "FF1D9E75"
DARK_FILL  = "FF111827"
BLUE_FILL  = "FF1E3A5F"
AMBER_FILL = "FFD97706"
TEAL_FILL  = "FF0E7490"
PINK_FILL  = "FFB91C1C"
GREY_FILL  = "FF4B5563"

WHITE_FONT = Font(color="FFFFFFFF", bold=True, name='Calibri', size=11)
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)


def _header_row(ws, cols, fill_hex):
    fill = PatternFill("solid", fgColor=fill_hex)
    for i, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i, value=col)
        cell.fill = fill
        cell.font = WHITE_FONT
        cell.alignment = HEADER_ALIGN
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def _auto_width(ws, cols):
    for i, col in enumerate(cols, 1):
        letter = get_column_letter(i)
        max_len = max(len(str(col)), 10)
        ws.column_dimensions[letter].width = min(max_len + 4, 40)


def _alt_row(ws, row_num):
    """Light grey fill for alternating rows."""
    if row_num % 2 == 0:
        fill = PatternFill("solid", fgColor="FFF3F4F6")
        for cell in ws[row_num]:
            cell.fill = fill


def export_excel(user_id, campana='2025/2026'):
    if not OPENPYXL:
        return ("openpyxl no instalado. Ejecuta: pip install openpyxl", 500)

    conn = get_db()
    wb = Workbook()

    # ── Explotación data ──
    explot = dicts(conn, "SELECT * FROM explotacion WHERE user_id=? LIMIT 1", (user_id,))
    ex = explot[0] if explot else {}
    titular = ex.get('titular', 'Explotación')
    municipio = ex.get('municipio', '')
    provincia = ex.get('provincia', '')
    now_str = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")

    # ══════════════════════════════════════════
    # HOJA 1 — PORTADA
    # ══════════════════════════════════════════
    ws = wb.active
    ws.title = "PORTADA"
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 45

    header_fill = PatternFill("solid", fgColor=GREEN_FILL)
    dark_fill   = PatternFill("solid", fgColor=DARK_FILL)

    ws.merge_cells('A1:B1')
    ws['A1'] = "CUADERNO DE CAMPO"
    ws['A1'].fill = dark_fill
    ws['A1'].font = Font(color="FFFFFFFF", bold=True, size=18, name='Calibri')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 40

    ws.merge_cells('A2:B2')
    ws['A2'] = "Cuaderno oficial de explotación agrícola"
    ws['A2'].fill = PatternFill("solid", fgColor=GREEN_FILL)
    ws['A2'].font = Font(color="FFFFFFFF", size=12, name='Calibri')
    ws['A2'].alignment = Alignment(horizontal='center')

    portada_data = [
        ("Titular", titular),
        ("NIF", ex.get('nif', '')),
        ("Municipio", municipio),
        ("Provincia", provincia),
        ("CP", ex.get('cp', '')),
        ("Teléfono", ex.get('telefono', '')),
        ("Email", ex.get('email', '')),
        ("Campaña", campana),
        ("Fecha de generación", now_str),
        ("RD aplicable", "RD 1311/2012"),
    ]
    for row_i, (label, val) in enumerate(portada_data, 4):
        c_label = ws.cell(row=row_i, column=1, value=label)
        c_label.font = Font(bold=True, name='Calibri')
        c_label.fill = PatternFill("solid", fgColor="FFE5E7EB")
        ws.cell(row=row_i, column=2, value=val)

    # ══════════════════════════════════════════
    # HOJA 2 — PARCELAS
    # ══════════════════════════════════════════
    ws2 = wb.create_sheet("PARCELAS")
    p_cols = ["ID", "Nombre Finca", "Polígono", "Parcela", "Recinto",
              "Provincia", "Municipio", "Uso SIGPAC", "Superficie (ha)",
              "Sistema Explotación", "Masa Agua <50m", "Notas"]
    _header_row(ws2, p_cols, GREEN_FILL)

    all_parcelas = dicts(conn, "SELECT * FROM parcelas WHERE user_id=? AND activa=1 ORDER BY nombre_finca", (user_id,))
    for ri, p in enumerate(all_parcelas, 2):
        row_data = [
            p.get('id'), p.get('nombre_finca'), p.get('poligono'),
            p.get('parcela_num'), p.get('recinto'), p.get('provincia_nombre'),
            p.get('municipio_nombre'), p.get('uso_sigpac'), p.get('superficie_ha'),
            p.get('sistema_explotacion'), "Sí" if p.get('masa_agua_cercana') else "No",
            p.get('notas'),
        ]
        for ci, val in enumerate(row_data, 1):
            ws2.cell(row=ri, column=ci, value=val)
        _alt_row(ws2, ri)
    _auto_width(ws2, p_cols)

    # ══════════════════════════════════════════
    # HOJA 3 — CULTIVOS POR CAMPAÑA
    # ══════════════════════════════════════════
    ws3 = wb.create_sheet("CULTIVOS POR CAMPAÑA")
    cc_cols = ["ID", "Parcela", "Campaña", "Cultivo", "Variedad",
               "Fecha Siembra", "Fecha Recol. Prevista", "Superficie Cultivada (ha)", "Notas"]
    _header_row(ws3, cc_cols, TEAL_FILL)
    cultivos = dicts(conn, """
        SELECT cc.*, p.nombre_finca FROM cultivos_campana cc
        LEFT JOIN parcelas p ON cc.parcela_id = p.id
        WHERE p.user_id=?
        ORDER BY cc.campana DESC, p.nombre_finca
    """, (user_id,))
    for ri, r in enumerate(cultivos, 2):
        row_data = [r.get('id'), r.get('nombre_finca'), r.get('campana'),
                    r.get('cultivo'), r.get('variedad'), r.get('fecha_siembra'),
                    r.get('fecha_recoleccion_prevista'), r.get('superficie_cultivada_ha'), r.get('notas')]
        for ci, val in enumerate(row_data, 1):
            ws3.cell(row=ri, column=ci, value=val)
        _alt_row(ws3, ri)
    _auto_width(ws3, cc_cols)

    # ══════════════════════════════════════════
    # HOJA 4 — TRATAMIENTOS FITOSANITARIOS
    # ══════════════════════════════════════════
    ws4 = wb.create_sheet("TRATAMIENTOS FITOSANITARIOS")
    t_cols = ["ID", "Parcela", "Fecha Aplicación", "Producto Comercial", "Nº Reg. MAPA",
              "Sustancia Activa", "Plaga/Objetivo", "Dosis", "Unidad", "Vol. Caldo (L/ha)",
              "Equipo", "Condic. Meteo.", "Plazo Seg. (días)", "Fecha Mín. Cosecha",
              "Eficacia", "Aplicador", "Notas", "Campaña"]
    _header_row(ws4, t_cols, GREEN_FILL)
    trats = dicts(conn, """
        SELECT t.*, p.nombre_finca, e.descripcion as equipo_nombre,
               a.nombre as aplicador_nombre
        FROM tratamientos t
        LEFT JOIN parcelas p ON t.parcela_id = p.id
        LEFT JOIN equipos e ON t.equipo_id = e.id
        LEFT JOIN aplicadores a ON t.aplicador_id = a.id
        WHERE t.user_id=? AND t.campana=?
        ORDER BY t.fecha_aplicacion DESC
    """, (user_id, campana))
    for ri, r in enumerate(trats, 2):
        row_data = [
            r.get('id'), r.get('nombre_finca') or r.get('parcela_etiqueta'),
            r.get('fecha_aplicacion'), r.get('producto_comercial'), r.get('num_registro_mapa'),
            r.get('sustancia_activa'), r.get('plaga_objetivo'),
            r.get('dosis_valor'), r.get('dosis_unidad'), r.get('volumen_caldo'),
            r.get('equipo_nombre'), r.get('condiciones_meteo'), r.get('plazo_seguridad_dias'),
            r.get('fecha_recoleccion_minima'), r.get('eficacia'), r.get('aplicador_nombre'),
            r.get('notas'), r.get('campana'),
        ]
        for ci, val in enumerate(row_data, 1):
            ws4.cell(row=ri, column=ci, value=val)
        _alt_row(ws4, ri)
    _auto_width(ws4, t_cols)

    # ══════════════════════════════════════════
    # HOJA 5 — FERTILIZACIÓN
    # ══════════════════════════════════════════
    ws5 = wb.create_sheet("FERTILIZACIÓN")
    f_cols = ["ID", "Parcela", "Fecha", "Tipo Fertilizante", "Producto",
              "Riqueza N-P-K", "Dosis", "Unidad", "Método", "Notas", "Campaña"]
    _header_row(ws5, f_cols, AMBER_FILL)
    fertz = dicts(conn, """
        SELECT f.*, p.nombre_finca FROM fertilizacion f
        LEFT JOIN parcelas p ON f.parcela_id = p.id
        WHERE f.user_id=? AND f.campana=?
        ORDER BY f.fecha_aplicacion DESC
    """, (user_id, campana))
    for ri, r in enumerate(fertz, 2):
        row_data = [r.get('id'), r.get('nombre_finca') or r.get('parcela_etiqueta'),
                    r.get('fecha_aplicacion'), r.get('tipo_fertilizante'), r.get('producto'),
                    r.get('riqueza_npk'), r.get('dosis_valor'), r.get('dosis_unidad'),
                    r.get('metodo_aplicacion'), r.get('notas'), r.get('campana')]
        for ci, val in enumerate(row_data, 1):
            ws5.cell(row=ri, column=ci, value=val)
        _alt_row(ws5, ri)
    _auto_width(ws5, f_cols)

    # ══════════════════════════════════════════
    # HOJA 6 — LABORES
    # ══════════════════════════════════════════
    ws6 = wb.create_sheet("LABORES")
    l_cols = ["ID", "Parcela", "Fecha", "Tipo Labor", "Descripción",
              "Maquinaria", "Horas", "Operario", "Notas", "Campaña"]
    _header_row(ws6, l_cols, BLUE_FILL)
    labores = dicts(conn, """
        SELECT l.*, p.nombre_finca FROM labores l
        LEFT JOIN parcelas p ON l.parcela_id = p.id
        WHERE l.user_id=? AND l.campana=?
        ORDER BY l.fecha DESC
    """, (user_id, campana))
    for ri, r in enumerate(labores, 2):
        row_data = [r.get('id'), r.get('nombre_finca') or r.get('parcela_etiqueta'),
                    r.get('fecha'), r.get('tipo_labor'), r.get('descripcion'),
                    r.get('maquinaria'), r.get('horas_trabajadas'), r.get('operario'),
                    r.get('notas'), r.get('campana')]
        for ci, val in enumerate(row_data, 1):
            ws6.cell(row=ri, column=ci, value=val)
        _alt_row(ws6, ri)
    _auto_width(ws6, l_cols)

    # ══════════════════════════════════════════
    # HOJA 7 — COSECHA
    # ══════════════════════════════════════════
    ws7 = wb.create_sheet("COSECHA")
    c_cols = ["ID", "Parcela", "Inicio Recol.", "Fin Recol.", "Cultivo",
              "Variedad", "Sup. Cosechada (ha)", "Producción Total", "Unidad",
              "Rendimiento (kg/ha)", "Destino", "Comprador", "Precio/Unidad", "Notas", "Campaña"]
    _header_row(ws7, c_cols, PINK_FILL)
    cosechas = dicts(conn, """
        SELECT c.*, p.nombre_finca FROM cosecha c
        LEFT JOIN parcelas p ON c.parcela_id = p.id
        WHERE c.user_id=? AND c.campana=?
        ORDER BY c.fecha_inicio DESC
    """, (user_id, campana))
    for ri, r in enumerate(cosechas, 2):
        row_data = [r.get('id'), r.get('nombre_finca') or r.get('parcela_etiqueta'),
                    r.get('fecha_inicio'), r.get('fecha_fin'), r.get('cultivo'),
                    r.get('variedad'), r.get('superficie_cosechada_ha'),
                    r.get('produccion_total_valor'), r.get('produccion_total_unidad'),
                    r.get('rendimiento_kg_ha'), r.get('destino'), r.get('comprador'),
                    r.get('precio_unidad'), r.get('notas'), r.get('campana')]
        for ci, val in enumerate(row_data, 1):
            ws7.cell(row=ri, column=ci, value=val)
        _alt_row(ws7, ri)
    _auto_width(ws7, c_cols)

    # ══════════════════════════════════════════
    # HOJA 8 — COMPRAS / VENTAS (solo si hay datos)
    # Obligatorio por RD 1311/2012 Anexo III Sección 5 cuando hay trazabilidad comercial
    # ══════════════════════════════════════════
    compras = dicts(conn, """
        SELECT * FROM compras
        WHERE user_id=? AND campana=? AND deleted_at IS NULL
        ORDER BY fecha ASC
    """, (user_id, campana))
    if compras:
        ws8 = wb.create_sheet("COMPRAS / VENTAS")
        cmp_cols = ["ID", "Fecha", "Tipo Producto", "Producto", "Nº Reg. MAPA",
                    "Sustancia Activa", "Proveedor", "Cantidad", "Unidad",
                    "Nº Lote", "Nº Factura", "Precio Total (€)", "Notas", "Campaña"]
        _header_row(ws8, cmp_cols, AMBER_FILL)
        for ri, r in enumerate(compras, 2):
            row_data = [
                r.get('id'), r.get('fecha'), r.get('tipo_producto'), r.get('producto'),
                r.get('num_registro_mapa'), r.get('sustancia_activa'), r.get('proveedor'),
                r.get('cantidad_valor'), r.get('cantidad_unidad'),
                r.get('num_lote'), r.get('num_factura'), r.get('precio_total'),
                r.get('notas'), r.get('campana'),
            ]
            for ci, val in enumerate(row_data, 1):
                ws8.cell(row=ri, column=ci, value=val)
            _alt_row(ws8, ri)
        _auto_width(ws8, cmp_cols)

    conn.close()

    # ── Save to BytesIO and send ──
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Cuaderno_Campo_{titular.replace(' ', '_')}_{campana.replace('/', '-')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
