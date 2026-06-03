import io
import json
import os
import re
import bcrypt
import requests as req_lib
import datetime
from functools import wraps
from flask import Flask, jsonify, request, send_file, session
from flask_cors import CORS
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_login import (LoginManager, UserMixin, login_user, logout_user,
                         login_required, current_user)
from db import (get_db, init_db, is_pac_eligible, dicts, one)
init_db()  # ejecutar siempre al importar (gunicorn + flask run)

frontend_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'frontend'))
app = Flask(__name__, static_folder=frontend_dir, static_url_path='')
_secret_key = os.environ.get('SECRET_KEY')
if not _secret_key:
    _is_prod = os.environ.get('FLASK_ENV') == 'production' or bool(os.environ.get('DATABASE_URL'))
    if _is_prod:
        raise RuntimeError("SECRET_KEY no está configurada. Establece la variable de entorno SECRET_KEY en producción.")
    import warnings
    warnings.warn("SECRET_KEY no configurada — usando clave de desarrollo insegura. NO usar en producción.")
    _secret_key = 'cuaderno_campo_DEV_ONLY_not_for_production'
app.secret_key = _secret_key

# CORS: en producción restringir a los orígenes del dominio real vía ALLOWED_ORIGINS
_allowed_origins = os.environ.get('ALLOWED_ORIGINS', 'http://127.0.0.1:5000,http://localhost:5000').split(',')
CORS(app, origins=_allowed_origins, supports_credentials=True)

# Rate limiting: máx 10 intentos de login por IP por minuto
limiter = Limiter(get_remote_address, app=app, default_limits=[],
                  storage_uri="memory://")

# ─────────────────────────────────────────────
# FLASK-LOGIN SETUP
# ─────────────────────────────────────────────
login_manager = LoginManager()
login_manager.init_app(app)

class User(UserMixin):
    def __init__(self, id, email, nombre, role, active, plan='trial', trial_ends_at=None, subscription_ends_at=None):
        self.id = id
        self.email = email
        self.nombre = nombre
        self.role = role
        self.active = active
        self.plan = plan
        self.trial_ends_at = trial_ends_at
        self.subscription_ends_at = subscription_ends_at

    def plan_is_active(self):
        """True si el usuario puede escribir datos (trial vigente, basic o pro)."""
        if self.role == 'admin':
            return True
        if self.plan in ('basic', 'pro'):
            return True
        if self.plan == 'trial' and self.trial_ends_at:
            ends = self.trial_ends_at
            if isinstance(ends, str):
                ends = datetime.datetime.fromisoformat(ends.replace('Z', ''))
            return datetime.datetime.utcnow() < ends
        return False

    def plan_label(self):
        """Estado legible para el frontend."""
        if self.plan in ('basic', 'pro'):
            return self.plan
        if self.plan == 'trial':
            if self.plan_is_active():
                return 'trial'
            return 'expired'
        return 'expired'

@login_manager.user_loader
def load_user(user_id):
    conn = get_db()
    u = one(conn, "SELECT * FROM users WHERE id=? AND active=1", (int(user_id),))
    conn.close()
    if not u:
        return None
    return User(u['id'], u['email'], u['nombre'], u['role'], u['active'],
                u.get('plan', 'trial'), u.get('trial_ends_at'), u.get('subscription_ends_at'))

@login_manager.unauthorized_handler
def unauthorized():
    return jsonify({"error": "No autenticado"}), 401

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'admin':
            return jsonify({"error": "No autorizado"}), 403
        return f(*args, **kwargs)
    return decorated

def get_uid():
    """Devuelve el user_id efectivo (admite impersonación del admin)."""
    if current_user.is_authenticated and current_user.role == 'admin':
        imp = session.get('impersonate_id')
        if imp:
            return imp
    return current_user.id

def requires_active_plan(f):
    """Decorador para rutas GET que también requieren suscripción activa (ej. exportaciones)."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.plan_is_active():
            return jsonify({"error": "subscription_required", "plan": current_user.plan_label()}), 403
        return f(*args, **kwargs)
    return decorated

_PLAN_EXEMPT_PREFIXES = ('/api/auth/', '/api/admin/', '/api/stripe/')

@app.before_request
def guard_active_plan():
    """Bloquea escrituras si el trial ha caducado o la suscripción ha expirado."""
    if request.method in ('GET', 'HEAD', 'OPTIONS'):
        return
    if any(request.path.startswith(p) for p in _PLAN_EXEMPT_PREFIXES):
        return
    if not current_user.is_authenticated:
        return
    if not current_user.plan_is_active():
        return jsonify({
            "error": "subscription_required",
            "plan": current_user.plan_label(),
        }), 403


# ─────────────────────────────────────────────
# DB HELPERS  (dicts / one imported from db.py)
# ─────────────────────────────────────────────


# ─────────────────────────────────────────────
# STATIC SERVING
# ─────────────────────────────────────────────
@app.route('/')
def serve_index():
    return app.send_static_file('index.html')

@app.route('/pago-completado')
def serve_pago_completado():
    return app.send_static_file('index.html')

@app.route('/<path:path>')
def serve_static(path):
    if path.startswith('api/'):
        return jsonify({"error": "Not found"}), 404
    try:
        return app.send_static_file(path)
    except Exception:
        return app.send_static_file('index.html')


# ─────────────────────────────────────────────
# AUTH
# ─────────────────────────────────────────────
@app.route('/api/auth/login', methods=['POST'])
@limiter.limit("10 per minute")
def auth_login():
    data = request.json or {}
    email = (data.get('email') or '').strip().lower()
    password = (data.get('password') or '').encode('utf-8')

    conn = get_db()
    u = one(conn, "SELECT * FROM users WHERE email=? AND active=1", (email,))
    conn.close()

    if not u or not bcrypt.checkpw(password, u['password_hash'].encode('utf-8')):
        return jsonify({"error": "Email o contraseña incorrectos"}), 401

    user = User(u['id'], u['email'], u['nombre'], u['role'], u['active'],
                u.get('plan', 'trial'), u.get('trial_ends_at'), u.get('subscription_ends_at'))
    login_user(user, remember=True)
    return jsonify({"id": user.id, "email": user.email,
                    "nombre": user.nombre, "role": user.role})


@app.route('/api/auth/register', methods=['POST'])
@limiter.limit("5 per minute")
def auth_register():
    data     = request.json or {}
    nombre   = (data.get('nombre') or '').strip()
    email    = (data.get('email') or '').strip().lower()
    password = (data.get('password') or '')

    if not nombre:
        return jsonify({"error": "El nombre es obligatorio"}), 400
    if not email or '@' not in email:
        return jsonify({"error": "Email no válido"}), 400
    if len(password) < 8:
        return jsonify({"error": "La contraseña debe tener al menos 8 caracteres"}), 400

    conn = get_db()
    existing = one(conn, "SELECT id FROM users WHERE email=?", (email,))
    if existing:
        conn.close()
        return jsonify({"error": "Ya existe una cuenta con ese email"}), 409

    pw_hash    = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    trial_ends = datetime.datetime.utcnow() + datetime.timedelta(days=7)
    try:
        c = conn.cursor()
        c.execute(
            "INSERT INTO users (email, password_hash, nombre, role, plan, trial_ends_at) VALUES (?,?,?,?,?,?)",
            (email, pw_hash, nombre, 'agricultor', 'trial', trial_ends.strftime('%Y-%m-%d %H:%M:%S'))
        )
        new_id = c.lastrowid
        c.execute("INSERT INTO explotacion (user_id, titular, campana_activa) VALUES (?,?,?)",
                  (new_id, nombre, '2025/2026'))
        conn.commit()
    except Exception as e:
        conn.close()
        app.logger.error("Register error: %s", e)
        return jsonify({"error": "Error al crear la cuenta"}), 500

    u = one(conn, "SELECT * FROM users WHERE id=?", (new_id,))
    conn.close()
    user = User(u['id'], u['email'], u['nombre'], u['role'], u['active'],
                u.get('plan', 'trial'), u.get('trial_ends_at'), u.get('subscription_ends_at'))
    login_user(user, remember=True)
    te = user.trial_ends_at
    return jsonify({
        "id": user.id, "email": user.email, "nombre": user.nombre, "role": user.role,
        "plan": user.plan_label(), "plan_raw": user.plan,
        "trial_ends_at": (te.isoformat() if hasattr(te, 'isoformat') else str(te)) if te else None,
        "plan_active": user.plan_is_active(), "impersonating": None,
    }), 201


@app.route('/api/auth/logout', methods=['POST'])
@login_required
def auth_logout():
    session.pop('impersonate_id', None)
    logout_user()
    return jsonify({"status": "ok"})

@app.route('/api/auth/me')
@login_required
def auth_me():
    imp_id = session.get('impersonate_id') if current_user.role == 'admin' else None
    imp_info = None
    if imp_id:
        conn = get_db()
        t = one(conn, "SELECT id, email, nombre FROM users WHERE id=?", (imp_id,))
        conn.close()
        if t:
            imp_info = t
    trial_ends = None
    if current_user.trial_ends_at:
        te = current_user.trial_ends_at
        trial_ends = te.isoformat() if hasattr(te, 'isoformat') else str(te)
    return jsonify({
        "id": current_user.id,
        "email": current_user.email,
        "nombre": current_user.nombre,
        "role": current_user.role,
        "impersonating": imp_info,
        "plan": current_user.plan_label(),
        "plan_raw": current_user.plan,
        "trial_ends_at": trial_ends,
        "plan_active": current_user.plan_is_active(),
    })

@app.route('/api/auth/change-password', methods=['POST'])
@login_required
def auth_change_password():
    data = request.json or {}
    old_pw = (data.get('old_password') or '').encode('utf-8')
    new_pw = (data.get('new_password') or '').encode('utf-8')
    if len(new_pw) < 8:
        return jsonify({"error": "La nueva contraseña debe tener al menos 8 caracteres"}), 400
    conn = get_db()
    u = one(conn, "SELECT * FROM users WHERE id=?", (current_user.id,))
    if not u or not bcrypt.checkpw(old_pw, u['password_hash'].encode('utf-8')):
        conn.close()
        return jsonify({"error": "Contraseña actual incorrecta"}), 401
    new_hash = bcrypt.hashpw(new_pw, bcrypt.gensalt()).decode('utf-8')
    conn.execute("UPDATE users SET password_hash=? WHERE id=?", (new_hash, current_user.id))
    conn.commit(); conn.close()
    return jsonify({"status": "ok"})


@app.route('/api/account/export-data', methods=['GET'])
@login_required
def account_export_data():
    """GDPR Art. 20 — portabilidad de datos. Devuelve todos los datos del usuario en JSON."""
    uid = current_user.id
    conn = get_db()
    export = {
        "usuario":         one(conn, "SELECT id,email,nombre,role,plan,created_at FROM users WHERE id=?", (uid,)),
        "explotacion":     dicts(conn, "SELECT * FROM explotacion WHERE user_id=?", (uid,)),
        "parcelas":        dicts(conn, "SELECT * FROM parcelas WHERE user_id=?", (uid,)),
        "tratamientos":    dicts(conn, "SELECT * FROM tratamientos WHERE user_id=? AND deleted_at IS NULL", (uid,)),
        "fertilizacion":   dicts(conn, "SELECT * FROM fertilizacion WHERE user_id=? AND deleted_at IS NULL", (uid,)),
        "abonado":         dicts(conn, "SELECT * FROM abonado WHERE user_id=? AND deleted_at IS NULL", (uid,)),
        "riego":           dicts(conn, "SELECT * FROM riego WHERE user_id=?", (uid,)),
        "labores":         dicts(conn, "SELECT * FROM labores WHERE user_id=?", (uid,)),
        "cosecha":         dicts(conn, "SELECT * FROM cosecha WHERE user_id=?", (uid,)),
        "compras":         dicts(conn, "SELECT * FROM compras WHERE user_id=? AND deleted_at IS NULL", (uid,)),
        "ventas":          dicts(conn, "SELECT * FROM ventas WHERE user_id=? AND deleted_at IS NULL", (uid,)),
        "cultivos_campana":dicts(conn, """SELECT cc.* FROM cultivos_campana cc
                                          JOIN parcelas p ON cc.parcela_id=p.id
                                          WHERE p.user_id=?""", (uid,)),
        "aplicadores":     dicts(conn, "SELECT * FROM aplicadores WHERE user_id=?", (uid,)),
        "equipos":         dicts(conn, "SELECT * FROM equipos WHERE user_id=?", (uid,)),
        "exportado_el":    datetime.datetime.utcnow().isoformat() + "Z",
        "normativa":       "RGPD Art. 20 — Derecho a la portabilidad de datos",
    }
    conn.close()
    buf = io.BytesIO(json.dumps(export, ensure_ascii=False, indent=2, default=str).encode('utf-8'))
    return send_file(buf, as_attachment=True,
                     download_name=f"datos_cuaderno_{uid}.json",
                     mimetype='application/json')


@app.route('/api/account/delete-account', methods=['DELETE'])
@login_required
def account_delete():
    """GDPR Art. 17 — derecho de supresión. Borra todos los datos del usuario y desactiva la cuenta."""
    uid = current_user.id
    conn = get_db()
    tables_soft = ['tratamientos', 'fertilizacion', 'abonado', 'compras', 'ventas']
    for t in tables_soft:
        conn.execute(f"UPDATE {t} SET deleted_at=datetime('now') WHERE user_id=? AND deleted_at IS NULL", (uid,))
    for t in ['labores', 'riego', 'cosecha', 'cultivos_campana', 'aplicadores', 'equipos']:
        try:
            conn.execute(f"DELETE FROM {t} WHERE user_id=?", (uid,))
        except Exception:
            pass
    conn.execute("UPDATE parcelas SET activa=0 WHERE user_id=?", (uid,))
    conn.execute("DELETE FROM explotacion WHERE user_id=?", (uid,))
    conn.execute(
        "UPDATE users SET active=0, email=?, nombre='[eliminado]', password_hash='', "
        "stripe_customer_id=NULL, stripe_subscription_id=NULL WHERE id=?",
        (f"deleted_{uid}_{int(datetime.datetime.utcnow().timestamp())}@borrado.invalid", uid)
    )
    conn.commit(); conn.close()
    logout_user()
    return jsonify({"status": "ok", "mensaje": "Cuenta y datos eliminados conforme al RGPD Art. 17"})


# ─────────────────────────────────────────────
# ADMIN — gestión de usuarios
# ─────────────────────────────────────────────
@app.route('/api/admin/users', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_users():
    conn = get_db()
    if request.method == 'GET':
        users = dicts(conn, "SELECT id,email,nombre,role,active,created_at FROM users ORDER BY created_at DESC")
        for u in users:
            uid = u['id']
            t = one(conn, "SELECT COUNT(*) as n FROM tratamientos WHERE user_id=?", (uid,))
            p = one(conn, "SELECT COUNT(*) as n FROM parcelas WHERE user_id=? AND activa=1", (uid,))
            l = one(conn, "SELECT COUNT(*) as n FROM labores WHERE user_id=?", (uid,))
            u['stats'] = {
                "tratamientos": t['n'] if t else 0,
                "parcelas": p['n'] if p else 0,
                "labores": l['n'] if l else 0,
            }
        conn.close()
        return jsonify(users)

    data = request.json or {}
    email = (data.get('email') or '').strip().lower()
    nombre = data.get('nombre') or ''
    password = data.get('password') or ''
    role = data.get('role', 'agricultor')

    if not email or not password:
        conn.close()
        return jsonify({"error": "Email y contraseña son obligatorios"}), 400

    pw_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    trial_ends = datetime.datetime.utcnow() + datetime.timedelta(days=7)
    try:
        c = conn.cursor()
        c.execute(
            "INSERT INTO users (email, password_hash, nombre, role, plan, trial_ends_at) VALUES (?,?,?,?,?,?)",
            (email, pw_hash, nombre, role, 'trial', trial_ends.strftime('%Y-%m-%d %H:%M:%S'))
        )
        new_id = c.lastrowid
        # Crear explotación vacía para el nuevo usuario
        c.execute("INSERT INTO explotacion (user_id, titular, campana_activa) VALUES (?,?,?)",
                  (new_id, nombre, '2025/2026'))
        conn.commit()
    except Exception as e:
        conn.close()
        if 'UNIQUE' in str(e):
            return jsonify({"error": "Ya existe un usuario con ese email"}), 409
        app.logger.error("Error creando usuario: %s", e)
        return jsonify({"error": "Error interno al crear el usuario"}), 500
    conn.close()
    return jsonify({"status": "ok", "id": new_id}), 201

@app.route('/api/admin/users/<int:uid>', methods=['PUT', 'DELETE'])
@login_required
@admin_required
def admin_user(uid):
    conn = get_db()
    if request.method == 'DELETE':
        conn.execute("UPDATE users SET active=0 WHERE id=?", (uid,))
        conn.commit(); conn.close()
        return jsonify({"status": "ok"})

    data = request.json or {}
    sets, vals = [], []
    if 'nombre' in data:
        sets.append('nombre=?'); vals.append(data['nombre'])
    if 'role' in data:
        sets.append('role=?'); vals.append(data['role'])
    if 'active' in data:
        sets.append('active=?'); vals.append(1 if data['active'] else 0)
    if data.get('password'):
        sets.append('password_hash=?')
        vals.append(bcrypt.hashpw(data['password'].encode('utf-8'), bcrypt.gensalt()).decode('utf-8'))
    if sets:
        conn.execute(f"UPDATE users SET {','.join(sets)} WHERE id=?", vals + [uid])
        conn.commit()
    conn.close()
    return jsonify({"status": "ok"})

@app.route('/api/admin/switch-user/<int:target_id>', methods=['POST'])
@login_required
@admin_required
def admin_switch_user(target_id):
    session['impersonate_id'] = target_id
    return jsonify({"status": "ok"})

@app.route('/api/admin/switch-back', methods=['POST'])
@login_required
@admin_required
def admin_switch_back():
    session.pop('impersonate_id', None)
    return jsonify({"status": "ok"})

@app.route('/api/admin/users/<int:uid>/export/pdf')
@login_required
@admin_required
def admin_export_pdf(uid):
    from export_pdf import export_pdf
    conn = get_db()
    row = conn.execute("SELECT id FROM users WHERE id=? AND active=1", (uid,)).fetchone()
    conn.close()
    if not row:
        return jsonify({"error": "Usuario no encontrado"}), 404
    campana = request.args.get('campana', '2025/2026')
    return export_pdf(uid, campana)


# ─────────────────────────────────────────────
# EXPLOTACION
# ─────────────────────────────────────────────
@app.route('/api/explotacion', methods=['GET', 'PUT', 'POST'])
@login_required
def explotacion():
    uid = get_uid()
    conn = get_db()
    if request.method == 'GET':
        row = one(conn, "SELECT * FROM explotacion WHERE user_id=? LIMIT 1", (uid,))
        # Admin sin explotación propia: usar la del agricultor principal (user_id=2)
        if not row and current_user.role == 'admin':
            row = one(conn, "SELECT * FROM explotacion WHERE user_id=2 LIMIT 1")
        conn.close()
        return jsonify(row or {})

    data = request.json or {}
    fields = ['titular', 'nif', 'municipio', 'provincia', 'cp',
              'telefono', 'email', 'campana_activa', 'fecha_apertura']
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM explotacion WHERE user_id=?", (uid,))
    if c.fetchone()[0] == 0:
        cols = ', '.join(['user_id'] + fields)
        vals = ', '.join(['?'] * (len(fields) + 1))
        c.execute(f"INSERT INTO explotacion ({cols}) VALUES ({vals})",
                  [uid] + [data.get(f) for f in fields])
    else:
        sets = ', '.join(f"{f}=?" for f in fields)
        c.execute(f"UPDATE explotacion SET {sets} WHERE user_id=?",
                  [data.get(f) for f in fields] + [uid])
    conn.commit(); conn.close()
    return jsonify({"status": "ok"})


# ─────────────────────────────────────────────
# STATS
# ─────────────────────────────────────────────
@app.route('/api/stats')
@login_required
def stats():
    uid = get_uid()
    conn = get_db()
    today = datetime.date.today().isoformat()
    next7 = (datetime.date.today() + datetime.timedelta(days=7)).isoformat()
    campana = request.args.get('campana', '2025/2026')

    all_p = dicts(conn, "SELECT uso_sigpac FROM parcelas WHERE user_id=? AND activa=1", (uid,))
    pac_count = sum(1 for p in all_p if is_pac_eligible(p['uso_sigpac']))

    t_count = one(conn, "SELECT COUNT(*) as n FROM tratamientos WHERE user_id=? AND campana=?", (uid, campana))
    f_count = one(conn, "SELECT COUNT(*) as n FROM fertilizacion WHERE user_id=? AND campana=?", (uid, campana))
    l_count = one(conn, "SELECT COUNT(*) as n FROM labores WHERE user_id=? AND campana=?", (uid, campana))
    c_count = one(conn, "SELECT COUNT(*) as n FROM cosecha WHERE user_id=? AND campana=?", (uid, campana))
    r_count = one(conn, "SELECT COUNT(*) as n FROM riego WHERE user_id=? AND campana=? AND deleted_at IS NULL", (uid, campana))
    a_count = one(conn, "SELECT COUNT(*) as n FROM abonado WHERE user_id=? AND campana=? AND deleted_at IS NULL", (uid, campana))

    alertas = dicts(conn, """
        SELECT parcela_etiqueta, producto_comercial, fecha_recoleccion_minima
        FROM tratamientos WHERE user_id=? AND fecha_recoleccion_minima >= ? AND fecha_recoleccion_minima <= ?
    """, (uid, today, next7))

    last_row = one(conn, """
        SELECT MAX(fecha) as last_fecha FROM (
            SELECT fecha_aplicacion as fecha FROM tratamientos WHERE user_id=?
            UNION ALL SELECT fecha FROM labores WHERE user_id=?
            UNION ALL SELECT fecha_aplicacion as fecha FROM fertilizacion WHERE user_id=?
            UNION ALL SELECT fecha FROM riego WHERE user_id=? AND deleted_at IS NULL
        )
    """, (uid, uid, uid, uid))
    days_inactive = 0
    if last_row and last_row.get('last_fecha'):
        try:
            last_date = datetime.date.fromisoformat(last_row['last_fecha'])
            days_inactive = (datetime.date.today() - last_date).days
        except Exception:
            days_inactive = 0

    conn.close()
    return jsonify({
        "parcelas_activas": pac_count,
        "tratamientos_mes": t_count['n'] if t_count else 0,
        "total_tratamientos": t_count['n'] if t_count else 0,
        "total_fertilizacion": f_count['n'] if f_count else 0,
        "total_labores": l_count['n'] if l_count else 0,
        "total_cosecha": c_count['n'] if c_count else 0,
        "total_riego": r_count['n'] if r_count else 0,
        "total_abonado": a_count['n'] if a_count else 0,
        "dias_sin_registro": days_inactive,
        "alertas_plazo": alertas,
    })


# ─────────────────────────────────────────────
# HISTORIAL
# ─────────────────────────────────────────────
@app.route('/api/historial')
@login_required
def historial():
    uid = get_uid()
    conn = get_db()
    parcela_id = request.args.get('parcela_id')
    modulo = request.args.get('modulo', 'todos')
    fecha_desde = request.args.get('fecha_desde', '')
    fecha_hasta = request.args.get('fecha_hasta', '')
    campana = request.args.get('campana', '')

    records = []

    def apply_filters(rows, date_field='fecha'):
        result = []
        for r in rows:
            if parcela_id and str(r.get('parcela_id')) != str(parcela_id):
                continue
            if campana and r.get('campana') != campana:
                continue
            f = r.get(date_field, '') or ''
            if fecha_desde and f < fecha_desde:
                continue
            if fecha_hasta and f > fecha_hasta:
                continue
            result.append(r)
        return result

    if modulo in ('todos', 'tratamientos'):
        rows = dicts(conn, "SELECT * FROM tratamientos WHERE user_id=? AND deleted_at IS NULL ORDER BY fecha_aplicacion DESC", (uid,))
        for r in apply_filters(rows, 'fecha_aplicacion'):
            records.append({**r, '_modulo': 'tratamientos', '_fecha': r.get('fecha_aplicacion', ''),
                            '_resumen': f"{r.get('producto_comercial','')} — {r.get('plaga_objetivo','')}"})

    if modulo in ('todos', 'fertilizacion'):
        rows = dicts(conn, "SELECT * FROM fertilizacion WHERE user_id=? AND deleted_at IS NULL ORDER BY fecha_aplicacion DESC", (uid,))
        for r in apply_filters(rows, 'fecha_aplicacion'):
            records.append({**r, '_modulo': 'fertilizacion', '_fecha': r.get('fecha_aplicacion', ''),
                            '_resumen': (
                                f"{r.get('tipo_fertilizante','')} — {r.get('producto','')}"
                                + (f" · N:{r['n_aplicado']} P:{r['p2o5_aplicado']} K:{r['k2o_aplicado']} kg/ha"
                                   if r.get('n_aplicado') is not None else '')
                            )})

    if modulo in ('todos', 'labores'):
        rows = dicts(conn, "SELECT * FROM labores WHERE user_id=? ORDER BY fecha DESC", (uid,))
        for r in apply_filters(rows, 'fecha'):
            desc = r.get('descripcion') or r.get('notas') or ''
            records.append({**r, '_modulo': 'labores', '_fecha': r.get('fecha', ''),
                            '_resumen': f"{r.get('tipo_labor','')} — {desc}".rstrip(' —')})

    if modulo in ('todos', 'cosecha'):
        rows = dicts(conn, "SELECT * FROM cosecha WHERE user_id=? ORDER BY fecha_inicio DESC", (uid,))
        for r in apply_filters(rows, 'fecha_inicio'):
            records.append({**r, '_modulo': 'cosecha', '_fecha': r.get('fecha_inicio', ''),
                            '_resumen': f"{r.get('cultivo','')} — {r.get('produccion_total_valor','')} {r.get('produccion_total_unidad','')}"})

    if modulo in ('todos', 'compras'):
        rows = dicts(conn, "SELECT * FROM compras WHERE user_id=? AND deleted_at IS NULL ORDER BY fecha DESC", (uid,))
        for r in apply_filters(rows, 'fecha'):
            records.append({**r, '_modulo': 'compras', '_fecha': r.get('fecha', ''),
                            '_resumen': f"{r.get('tipo_producto','')} — {r.get('producto','')} · {r.get('proveedor','')}"})

    if modulo in ('todos', 'riego'):
        rows = dicts(conn, "SELECT * FROM riego WHERE user_id=? AND deleted_at IS NULL ORDER BY fecha DESC", (uid,))
        for r in apply_filters(rows, 'fecha'):
            vol = f"{r['volumen_m3']} m³" if r.get('volumen_m3') else f"{r.get('horas_riego','')} h"
            records.append({**r, '_modulo': 'riego', '_fecha': r.get('fecha', ''),
                            '_resumen': f"{r.get('tipo_riego','')} — {vol}"})

    if modulo in ('todos', 'abonado'):
        rows = dicts(conn, "SELECT * FROM abonado WHERE user_id=? AND deleted_at IS NULL ORDER BY fecha_preparacion DESC", (uid,))
        for r in apply_filters(rows, 'fecha_preparacion'):
            records.append({**r, '_modulo': 'abonado', '_fecha': r.get('fecha_preparacion', ''),
                            '_resumen': f"{r.get('cultivo','')} — N:{r.get('n_necesario_kg_ha','')} P:{r.get('p_necesario_kg_ha','')} K:{r.get('k_necesario_kg_ha','')} kg/ha"})

    records.sort(key=lambda x: x.get('_fecha', '') or '', reverse=True)
    conn.close()
    return jsonify(records)


# ─────────────────────────────────────────────
# PARCELAS
# ─────────────────────────────────────────────
@app.route('/api/parcelas', methods=['GET', 'POST'])
@login_required
def manage_parcelas():
    uid = get_uid()
    conn = get_db()

    if request.method == 'GET':
        all_p = dicts(conn, "SELECT * FROM parcelas WHERE user_id=? AND activa=1 ORDER BY nombre_finca", (uid,))
        pac_only = request.args.get('pac_only', 'false').lower() == 'true'
        if pac_only:
            all_p = [p for p in all_p if is_pac_eligible(p.get('uso_sigpac', ''))]
        conn.close()
        return jsonify(all_p)

    data = request.json or {}

    def _to_float(v):
        if v is None or v == '': return None
        try: return float(v)
        except (ValueError, TypeError): return None

    sup = _to_float(data.get('superficie_ha'))
    if sup is not None and sup <= 0:
        conn.close()
        return jsonify({"error": "La superficie debe ser mayor que cero"}), 400

    c = conn.cursor()
    c.execute('''
        INSERT INTO parcelas (
            user_id, comunidad, provincia_cod, provincia_nombre,
            municipio_cod, municipio_nombre, nombre_finca,
            poligono, parcela_num, recinto, superficie_ha, uso_sigpac,
            sistema_explotacion, masa_agua_cercana, notas
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    ''', (
        uid, data.get('comunidad'), data.get('provincia_cod'), data.get('provincia_nombre'),
        data.get('municipio_cod'), data.get('municipio_nombre'), data.get('nombre_finca'),
        data.get('poligono'), data.get('parcela_num'), data.get('recinto'),
        sup, data.get('uso_sigpac'),
        data.get('sistema_explotacion', 'Secano'),
        1 if data.get('masa_agua_cercana') else 0,
        data.get('notas'),
    ))
    new_id = c.lastrowid
    conn.commit(); conn.close()
    return jsonify({"status": "ok", "id": new_id}), 201


@app.route('/api/parcelas/<int:pid>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def manage_parcela(pid):
    uid = get_uid()
    conn = get_db()
    if request.method == 'GET':
        row = one(conn, "SELECT * FROM parcelas WHERE id=? AND user_id=?", (pid, uid))
        conn.close()
        return jsonify(row or {})

    if request.method == 'DELETE':
        orphan_counts = {}
        for table, label in [('tratamientos', 'tratamientos'), ('fertilizacion', 'fertilizaciones'),
                              ('cosecha', 'cosechas'), ('labores', 'labores'), ('riego', 'riegos')]:
            try:
                row = one(conn, f"SELECT COUNT(*) as n FROM {table} WHERE parcela_id=? AND user_id=?", (pid, uid))
                if row and row['n']:
                    orphan_counts[label] = row['n']
            except Exception:
                pass
        conn.execute("UPDATE parcelas SET activa=0 WHERE id=? AND user_id=?", (pid, uid))
        conn.commit(); conn.close()
        resp = {"status": "ok"}
        if orphan_counts:
            detalle = ', '.join(f"{n} {k}" for k, n in orphan_counts.items())
            resp["warning"] = f"La parcela tenía registros asociados: {detalle}. Siguen en el historial pero sin parcela activa."
        return jsonify(resp)

    data = request.json or {}

    def _to_float(v):
        if v is None or v == '': return None
        try: return float(v)
        except (ValueError, TypeError): return None

    sup_put = _to_float(data.get('superficie_ha'))
    if sup_put is not None and sup_put <= 0:
        conn.close()
        return jsonify({"error": "La superficie debe ser mayor que cero"}), 400

    def _field_val(f):
        v = data.get(f)
        if f == 'superficie_ha': return sup_put
        if f == 'masa_agua_cercana': return 1 if v else 0
        return v

    fields = ['comunidad','provincia_cod','provincia_nombre','municipio_cod','municipio_nombre',
              'nombre_finca','poligono','parcela_num','recinto','superficie_ha','uso_sigpac',
              'sistema_explotacion','masa_agua_cercana','notas']
    sets = ', '.join(f"{f}=?" for f in fields)
    vals = [_field_val(f) for f in fields] + [pid, uid]
    conn.execute(f"UPDATE parcelas SET {sets} WHERE id=? AND user_id=?", vals)
    conn.commit(); conn.close()
    return jsonify({"status": "ok"})


# ─────────────────────────────────────────────
# CULTIVOS CAMPAÑA
# ─────────────────────────────────────────────
@app.route('/api/cultivos-campana', methods=['GET', 'POST'])
@login_required
def manage_cultivos():
    uid = get_uid()
    conn = get_db()
    if request.method == 'GET':
        parcela_id = request.args.get('parcela_id')
        campana = request.args.get('campana')
        # Filtrar siempre por user_id a través de la parcela propietaria
        sql = """SELECT cc.* FROM cultivos_campana cc
                 JOIN parcelas p ON cc.parcela_id = p.id
                 WHERE p.user_id=?"""
        params = [uid]
        if parcela_id:
            sql += " AND cc.parcela_id=?"; params.append(parcela_id)
        if campana:
            sql += " AND cc.campana=?"; params.append(campana)
        rows = dicts(conn, sql, params)
        conn.close()
        return jsonify(rows)

    data = request.json or {}
    # Verificar que la parcela pertenece al usuario
    parcela_id = data.get('parcela_id')
    if not parcela_id:
        conn.close()
        return jsonify({"error": "Parcela es obligatoria"}), 400
    owner = one(conn, "SELECT id FROM parcelas WHERE id=? AND user_id=?", (parcela_id, uid))
    if not owner:
        conn.close()
        return jsonify({"error": "Parcela no encontrada"}), 404
    if not data.get('cultivo'):
        conn.close()
        return jsonify({"error": "El cultivo es obligatorio"}), 400
    if not data.get('cultivo_iacs_cod'):
        conn.close()
        return jsonify({"error": "El código IACS del cultivo es obligatorio para la interoperabilidad con SIEX (obligatorio desde ene 2027)"}), 400
    c = conn.cursor()
    try:
        c.execute('''
            INSERT INTO cultivos_campana
                (parcela_id, campana, cultivo, cultivo_iacs_cod, variedad, fecha_siembra,
                 fecha_recoleccion_prevista, superficie_cultivada_ha, notas)
            VALUES (?,?,?,?,?,?,?,?,?)
        ''', (parcela_id, data.get('campana'), data.get('cultivo'),
              data.get('cultivo_iacs_cod'),
              data.get('variedad'), data.get('fecha_siembra'),
              data.get('fecha_recoleccion_prevista'), data.get('superficie_cultivada_ha'),
              data.get('notas')))
        new_id = c.lastrowid
    except Exception:
        c.execute('''
            UPDATE cultivos_campana SET cultivo=?, cultivo_iacs_cod=?, variedad=?,
                fecha_siembra=?, fecha_recoleccion_prevista=?,
                superficie_cultivada_ha=?, notas=?, updated_at=CURRENT_TIMESTAMP
            WHERE parcela_id=? AND campana=?
        ''', (data.get('cultivo'), data.get('cultivo_iacs_cod'),
              data.get('variedad'), data.get('fecha_siembra'),
              data.get('fecha_recoleccion_prevista'), data.get('superficie_cultivada_ha'),
              data.get('notas'), parcela_id, data.get('campana')))
        new_id = None
    conn.commit(); conn.close()
    return jsonify({"status": "ok", "id": new_id}), 201


@app.route('/api/cultivos-campana/<int:cid>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def manage_cultivo(cid):
    uid = get_uid()
    conn = get_db()
    # Verificar propiedad a través de la parcela (cultivos_campana no tiene user_id propio)
    owner = one(conn, """SELECT cc.id FROM cultivos_campana cc
                         JOIN parcelas p ON cc.parcela_id = p.id
                         WHERE cc.id=? AND p.user_id=?""", (cid, uid))
    if not owner:
        conn.close()
        return jsonify({"error": "No encontrado"}), 404
    if request.method == 'DELETE':
        conn.execute("DELETE FROM cultivos_campana WHERE id=?", (cid,))
        conn.commit(); conn.close()
        return jsonify({"status": "ok"})
    if request.method == 'GET':
        row = one(conn, "SELECT * FROM cultivos_campana WHERE id=?", (cid,))
        conn.close()
        return jsonify(row or {})
    data = request.json or {}
    fields = ['cultivo','cultivo_iacs_cod','variedad','fecha_siembra','fecha_recoleccion_prevista','superficie_cultivada_ha','notas']
    sets = ', '.join(f"{f}=?" for f in fields)
    conn.execute(f"UPDATE cultivos_campana SET {sets} WHERE id=?",
                 [data.get(f) for f in fields] + [cid])
    conn.commit(); conn.close()
    return jsonify({"status": "ok"})


# ─────────────────────────────────────────────
# TRATAMIENTOS
# ─────────────────────────────────────────────
# VALIDADORES RD 1311/2012
# ─────────────────────────────────────────────
def _calc_fecha_recoleccion(fecha_aplicacion_str, plazo_dias):
    """Calcula fecha mínima de recolección. Siempre calculada en backend, nunca confiamos en el cliente."""
    try:
        fecha = datetime.date.fromisoformat(str(fecha_aplicacion_str))
        plazo = int(plazo_dias or 0)
        return (fecha + datetime.timedelta(days=plazo)).isoformat()
    except (ValueError, TypeError):
        return None


def _validate_tratamiento(data):
    """Devuelve mensaje de error si faltan campos obligatorios (Anexo III S3)."""
    required = {
        'fecha_aplicacion':    'Fecha de aplicación',
        'parcela_id':          'Parcela SIGPAC (Anexo III S3)',
        'producto_comercial':  'Producto comercial',
        'num_registro_mapa':   'Nº Registro MAPA',
        'sustancia_activa':    'Sustancia activa',
        'plaga_objetivo':      'Plaga / enfermedad objetivo',
        'dosis_valor':         'Dosis',
        'aplicador_id':        'Aplicador (obligatorio por ROPO)',
        'equipo_aplicacion':   'Equipo de aplicación (Anexo III S3)',
        'plazo_seguridad_dias':'Plazo de seguridad (días)',
    }
    missing = [label for field, label in required.items() if not data.get(field) and data.get(field) != 0]
    if missing:
        return f"Campos obligatorios según RD 1311/2012: {', '.join(missing)}"
    try:
        fecha = datetime.date.fromisoformat(str(data['fecha_aplicacion']))
        if fecha > datetime.date.today():
            return "La fecha de aplicación no puede ser futura"
    except (ValueError, TypeError):
        return "Fecha de aplicación con formato inválido (use YYYY-MM-DD)"
    try:
        if int(data['plazo_seguridad_dias']) < 0:
            return "El plazo de seguridad no puede ser negativo"
    except (ValueError, TypeError):
        return "El plazo de seguridad debe ser un número entero"
    try:
        if float(data['dosis_valor']) <= 0:
            return "La dosis debe ser mayor que cero"
    except (ValueError, TypeError):
        return "La dosis debe ser un número válido"
    mapa = str(data.get('num_registro_mapa', '')).strip()
    if not re.fullmatch(r'\d{4,6}(/\d+)?', mapa):
        return "El Nº de Registro MAPA debe ser numérico (ej: 12345 o 12345/2)"
    return None


def _validate_campana(campana):
    """Valida que la campaña tenga formato YYYY/YYYY con años consecutivos."""
    if not campana:
        return None
    if not re.fullmatch(r'\d{4}/\d{4}', str(campana)):
        return "El campo campaña debe tener formato YYYY/YYYY (ej: 2025/2026)"
    y1, y2 = int(str(campana)[:4]), int(str(campana)[5:])
    if y2 != y1 + 1:
        return "La campaña debe ser de años consecutivos (ej: 2025/2026)"
    return None


def _validate_compra(data):
    """Valida campos obligatorios de compras (Anexo III S5 — trazabilidad)."""
    required = {
        'fecha':         'Fecha de compra',
        'tipo_producto': 'Tipo de producto',
        'producto':      'Nombre del producto',
    }
    missing = [label for field, label in required.items() if not data.get(field)]
    if missing:
        return f"Campos obligatorios: {', '.join(missing)}"
    if data.get('tipo_producto') == 'fitosanitario':
        fito_required = {
            'num_registro_mapa': 'Nº de registro MAPA',
            'sustancia_activa':  'Sustancia activa',
        }
        missing_fito = [label for field, label in fito_required.items() if not data.get(field)]
        if missing_fito:
            return f"Obligatorio para fitosanitarios (RD 1311/2012 Anexo III S5): {', '.join(missing_fito)}"
    try:
        fecha = datetime.date.fromisoformat(str(data['fecha']))
        if fecha > datetime.date.today():
            return "La fecha de compra no puede ser futura"
    except (ValueError, TypeError):
        return "Fecha con formato inválido (use YYYY-MM-DD)"
    return None


def _validate_fertilizacion(data):
    """Devuelve mensaje de error si faltan campos obligatorios (Anexo III S4)."""
    required = {
        'fecha_aplicacion': 'Fecha de aplicación',
        'parcela_id':       'Parcela SIGPAC (Anexo III S4)',
        'tipo_fertilizante': 'Tipo de fertilizante',
        'producto':          'Nombre del producto',
        'dosis_valor':       'Dosis (cantidad)',
    }
    missing = [label for field, label in required.items() if not data.get(field) and data.get(field) != 0]
    if missing:
        return f"Campos obligatorios según RD 1311/2012: {', '.join(missing)}"
    try:
        fecha = datetime.date.fromisoformat(str(data['fecha_aplicacion']))
        if fecha > datetime.date.today():
            return "La fecha de aplicación no puede ser futura"
    except (ValueError, TypeError):
        return "Fecha de aplicación con formato inválido (use YYYY-MM-DD)"
    try:
        if float(data['dosis_valor']) <= 0:
            return "La dosis debe ser mayor que cero"
    except (ValueError, TypeError):
        return "La dosis debe ser un número válido"
    if str(data.get('dosis_unidad', '')).strip().lower() in ('l/ha', 'l ha', 'litros/ha'):
        dens = data.get('densidad_g_ml')
        if not dens:
            return "Para fertilizantes líquidos (L/ha) indica la densidad en g/mL para calcular NPK correctamente"
        try:
            if float(str(dens).replace(',', '.')) <= 0:
                return "La densidad debe ser mayor que cero"
        except (ValueError, TypeError):
            return "La densidad debe ser un número válido (g/mL)"
    return None


def _calc_npk(riqueza_npk, dosis_valor, dosis_unidad='kg/ha', densidad_g_ml=None):
    """Parsea riqueza N-P-K y devuelve (n, p, k) en kg/ha.
    Soporta: '15-15-15', '27-0-0', '46%N', '34,5%N', '18%N 46%P2O5 0%K2O'.
    Para líquidos (L/ha) requiere densidad_g_ml para convertir a kg/ha correctamente.
    Devuelve (None, None, None) si no es parseable."""
    if not riqueza_npk or not dosis_valor:
        return None, None, None
    try:
        dosis = float(str(dosis_valor).replace(',', '.'))
        if dosis <= 0:
            return None, None, None
    except (ValueError, TypeError):
        return None, None, None
    # Convertir L/ha → kg/ha usando densidad (g/mL = kg/L)
    if str(dosis_unidad).strip().lower() in ('l/ha', 'l ha', 'litros/ha'):
        try:
            dens = float(str(densidad_g_ml).replace(',', '.'))
            if dens > 0:
                dosis = dosis * dens
            else:
                return None, None, None
        except (ValueError, TypeError):
            return None, None, None
    s = str(riqueza_npk).replace(',', '.').upper()
    # Formato triple N-P-K (ej: "15-15-15", "27-0-0", "15 15 15")
    m3 = re.search(r'(\d+\.?\d*)[^\d]+(\d+\.?\d*)[^\d]+(\d+\.?\d*)', s)
    if m3:
        try:
            return (round(float(m3.group(1)) / 100 * dosis, 2),
                    round(float(m3.group(2)) / 100 * dosis, 2),
                    round(float(m3.group(3)) / 100 * dosis, 2))
        except (ValueError, TypeError):
            pass
    # Formato con letra: "46%N", "34.5%N 0%P 0%K", "18%N46%P2O5"
    def _pct(letter):
        m = re.search(r'(\d+\.?\d*)\s*%?\s*' + letter + r'(?:[^A-Z]|$)', s)
        if m:
            try:
                return round(float(m.group(1)) / 100 * dosis, 2)
            except (ValueError, TypeError):
                pass
        return None
    n = _pct('N')
    p = _pct('P')
    k = _pct('K')
    if any(v is not None for v in (n, p, k)):
        return (n or 0.0, p or 0.0, k or 0.0)
    return None, None, None


def _validate_riego(data):
    """Valida campos obligatorios del registro de riego (RD 934/2025).
    Requiere fecha, parcela, tipo y al menos horas o m³."""
    required = {
        'fecha':      'Fecha de riego',
        'parcela_id': 'Parcela',
        'tipo_riego': 'Tipo de riego',
    }
    missing = [label for field, label in required.items() if not data.get(field)]
    if missing:
        return f"Campos obligatorios: {', '.join(missing)}"
    if not data.get('horas_riego') and not data.get('volumen_m3'):
        return "Indica al menos las horas de riego o el volumen en m³"
    try:
        fecha = datetime.date.fromisoformat(str(data['fecha']))
        if fecha > datetime.date.today():
            return "La fecha de riego no puede ser futura"
    except (ValueError, TypeError):
        return "Fecha con formato inválido (use YYYY-MM-DD)"
    return None


def _validate_abonado(data):
    required = {
        'parcela_id':               'Parcela',
        'cultivo':                  'Cultivo',
        'cultivo_anterior':         'Cultivo anterior',
        'rendimiento_esperado_kg_ha': 'Rendimiento esperado',
        'fecha_preparacion':        'Fecha de preparación',
        'datos_suelo':              'Datos de análisis de suelo (RD 934/2025)',
        'n_necesario_kg_ha':        'N necesario',
        'p_necesario_kg_ha':        'P necesario',
        'k_necesario_kg_ha':        'K necesario',
    }
    missing = [label for field, label in required.items() if not str(data.get(field, '')).strip()]
    if missing:
        return f"Campos obligatorios: {', '.join(missing)}"
    try:
        fecha = datetime.date.fromisoformat(str(data['fecha_preparacion']))
        if fecha > datetime.date.today():
            return "La fecha de preparación no puede ser futura"
    except (ValueError, TypeError):
        return "Fecha con formato inválido (use YYYY-MM-DD)"
    return None


# ─────────────────────────────────────────────
@app.route('/api/tratamientos', methods=['GET', 'POST'])
@login_required
def manage_tratamientos():
    uid = get_uid()
    conn = get_db()
    if request.method == 'GET':
        rows = dicts(conn, "SELECT * FROM tratamientos WHERE user_id=? AND deleted_at IS NULL ORDER BY fecha_aplicacion DESC", (uid,))
        conn.close()
        return jsonify(rows)

    data = request.json or {}
    err = _validate_tratamiento(data) or _validate_campana(data.get('campana'))
    if err:
        conn.close()
        return jsonify({"error": err}), 400

    # Verificar que el aplicador seleccionado tiene ROPO registrado (RD 1311/2012)
    if data.get('aplicador_id'):
        aplicador = one(conn, "SELECT num_ropo FROM aplicadores WHERE id=? AND user_id=?",
                        (data['aplicador_id'], uid))
        if not aplicador or not (aplicador.get('num_ropo') or '').strip():
            conn.close()
            return jsonify({"error": (
                "El aplicador seleccionado no tiene número ROPO registrado. "
                "El ROPO es obligatorio según el RD 1311/2012 Anexo III S3. "
                "Edita el aplicador y añade su número de carnet ROPO."
            )}), 400

    c = conn.cursor()
    c.execute('''
        INSERT INTO tratamientos (
            user_id, parcela_id, parcela_etiqueta, fecha_aplicacion,
            producto_comercial, num_registro_mapa, sustancia_activa,
            plaga_objetivo, dosis_valor, dosis_unidad, volumen_caldo,
            equipo_id, condiciones_meteo, plazo_seguridad_dias,
            fecha_recoleccion_minima, eficacia, aplicador_id, notas, campana
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    ''', (
        uid, data.get('parcela_id'), data.get('parcela_etiqueta'), data.get('fecha_aplicacion'),
        data.get('producto_comercial'), data.get('num_registro_mapa'), data.get('sustancia_activa'),
        data.get('plaga_objetivo'), data.get('dosis_valor') or None, data.get('dosis_unidad', 'L/ha'),
        data.get('volumen_caldo') or None, data.get('equipo_id') or None, data.get('condiciones_meteo'),
        data.get('plazo_seguridad_dias') or None,
        _calc_fecha_recoleccion(data.get('fecha_aplicacion'), data.get('plazo_seguridad_dias')),
        data.get('eficacia'), data.get('aplicador_id') or None, data.get('notas'),
        data.get('campana', '2025/2026'),
    ))
    conn.commit(); new_id = c.lastrowid; conn.close()
    return jsonify({"status": "ok", "id": new_id}), 201


@app.route('/api/tratamientos/<int:tid>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def manage_tratamiento(tid):
    uid = get_uid()
    conn = get_db()
    if request.method == 'DELETE':
        conn.execute(
            "UPDATE tratamientos SET deleted_at=datetime('now') WHERE id=? AND user_id=?",
            (tid, uid))
        conn.commit(); conn.close(); return jsonify({"status": "ok"})
    if request.method == 'GET':
        row = one(conn, "SELECT * FROM tratamientos WHERE id=? AND user_id=? AND deleted_at IS NULL", (tid, uid))
        conn.close(); return jsonify(row or {})
    data = request.json or {}
    err = _validate_tratamiento(data)
    if err:
        conn.close()
        return jsonify({"error": err}), 400
    if data.get('aplicador_id'):
        aplicador = one(conn, "SELECT num_ropo FROM aplicadores WHERE id=? AND user_id=?",
                        (data['aplicador_id'], uid))
        if not aplicador or not (aplicador.get('num_ropo') or '').strip():
            conn.close()
            return jsonify({"error": (
                "El aplicador seleccionado no tiene número ROPO registrado. "
                "Edita el aplicador y añade su número de carnet ROPO antes de guardar."
            )}), 400
    # Calcular siempre en backend, nunca confiar en el cliente
    data['fecha_recoleccion_minima'] = _calc_fecha_recoleccion(
        data.get('fecha_aplicacion'), data.get('plazo_seguridad_dias'))
    fields = ['parcela_id','parcela_etiqueta','fecha_aplicacion','producto_comercial',
              'num_registro_mapa','sustancia_activa','plaga_objetivo','dosis_valor','dosis_unidad',
              'volumen_caldo','equipo_id','condiciones_meteo','plazo_seguridad_dias',
              'fecha_recoleccion_minima','eficacia','aplicador_id','notas','campana']
    sets = ', '.join(f"{f}=?" for f in fields)
    _numeric_t = {'dosis_valor', 'volumen_caldo', 'equipo_id', 'plazo_seguridad_dias', 'aplicador_id'}
    conn.execute(f"UPDATE tratamientos SET {sets} WHERE id=? AND user_id=? AND deleted_at IS NULL",
                 [data.get(f) or None if f in _numeric_t else data.get(f) for f in fields] + [tid, uid])
    conn.commit(); conn.close(); return jsonify({"status": "ok"})


# ─────────────────────────────────────────────
# FERTILIZACIÓN
# ─────────────────────────────────────────────
@app.route('/api/fertilizacion', methods=['GET', 'POST'])
@login_required
def manage_fertilizacion():
    uid = get_uid()
    conn = get_db()
    if request.method == 'GET':
        rows = dicts(conn, "SELECT * FROM fertilizacion WHERE user_id=? AND deleted_at IS NULL ORDER BY fecha_aplicacion DESC", (uid,))
        conn.close(); return jsonify(rows)

    data = request.json or {}
    err = _validate_fertilizacion(data) or _validate_campana(data.get('campana'))
    if err:
        conn.close()
        return jsonify({"error": err}), 400

    n_ap, p_ap, k_ap = _calc_npk(data.get('riqueza_npk'), data.get('dosis_valor'),
                                  data.get('dosis_unidad', 'kg/ha'), data.get('densidad_g_ml'))
    c = conn.cursor()
    c.execute('''
        INSERT INTO fertilizacion (
            user_id, parcela_id, parcela_etiqueta, fecha_aplicacion,
            tipo_fertilizante, producto, riqueza_npk,
            dosis_valor, dosis_unidad, densidad_g_ml, metodo_aplicacion, notas, campana,
            n_aplicado, p2o5_aplicado, k2o_aplicado
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    ''', (uid, data.get('parcela_id'), data.get('parcela_etiqueta'), data.get('fecha_aplicacion'),
          data.get('tipo_fertilizante'), data.get('producto'), data.get('riqueza_npk'),
          data.get('dosis_valor') or None, data.get('dosis_unidad', 'kg/ha'),
          data.get('densidad_g_ml') or None,
          data.get('metodo_aplicacion'), data.get('notas'), data.get('campana', '2025/2026'),
          n_ap, p_ap, k_ap))
    conn.commit(); new_id = c.lastrowid; conn.close()
    return jsonify({"status": "ok", "id": new_id}), 201


@app.route('/api/fertilizacion/<int:fid>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def manage_fertilizacion_one(fid):
    uid = get_uid()
    conn = get_db()
    if request.method == 'DELETE':
        conn.execute(
            "UPDATE fertilizacion SET deleted_at=datetime('now') WHERE id=? AND user_id=?",
            (fid, uid))
        conn.commit(); conn.close(); return jsonify({"status": "ok"})
    if request.method == 'GET':
        row = one(conn, "SELECT * FROM fertilizacion WHERE id=? AND user_id=? AND deleted_at IS NULL", (fid, uid))
        conn.close(); return jsonify(row or {})
    data = request.json or {}
    err = _validate_fertilizacion(data)
    if err:
        conn.close()
        return jsonify({"error": err}), 400
    n_ap, p_ap, k_ap = _calc_npk(data.get('riqueza_npk'), data.get('dosis_valor'),
                                  data.get('dosis_unidad', 'kg/ha'), data.get('densidad_g_ml'))
    fields = ['parcela_id','parcela_etiqueta','fecha_aplicacion','tipo_fertilizante',
              'producto','riqueza_npk','dosis_valor','dosis_unidad','densidad_g_ml',
              'metodo_aplicacion','notas','campana',
              'n_aplicado','p2o5_aplicado','k2o_aplicado']
    sets = ', '.join(f"{f}=?" for f in fields)
    _numeric_f = {'dosis_valor', 'densidad_g_ml'}
    npk_map = {'n_aplicado': n_ap, 'p2o5_aplicado': p_ap, 'k2o_aplicado': k_ap}
    values = [data.get(f) or None if f in _numeric_f else npk_map.get(f, data.get(f)) for f in fields]
    conn.execute(f"UPDATE fertilizacion SET {sets} WHERE id=? AND user_id=? AND deleted_at IS NULL",
                 values + [fid, uid])
    conn.commit(); conn.close(); return jsonify({"status": "ok"})


# ─────────────────────────────────────────────
# RIEGO
# ─────────────────────────────────────────────
@app.route('/api/riego', methods=['GET', 'POST'])
@login_required
def manage_riego():
    uid = get_uid()
    conn = get_db()
    if request.method == 'GET':
        campana = request.args.get('campana', '2025/2026')
        rows = dicts(conn, "SELECT * FROM riego WHERE user_id=? AND campana=? AND deleted_at IS NULL ORDER BY fecha DESC", (uid, campana))
        conn.close(); return jsonify(rows)

    data = request.json or {}
    err = _validate_riego(data)
    if err:
        conn.close()
        return jsonify({"error": err}), 400

    c = conn.cursor()
    c.execute('''
        INSERT INTO riego (
            user_id, parcela_id, parcela_etiqueta, fecha,
            tipo_riego, volumen_m3, horas_riego, fuente_agua, notas, campana
        ) VALUES (?,?,?,?,?,?,?,?,?,?)
    ''', (uid, data.get('parcela_id'), data.get('parcela_etiqueta'), data.get('fecha'),
          data.get('tipo_riego'), data.get('volumen_m3') or None,
          data.get('horas_riego') or None, data.get('fuente_agua'), data.get('notas'),
          data.get('campana', '2025/2026')))
    conn.commit(); new_id = c.lastrowid; conn.close()
    return jsonify({"status": "ok", "id": new_id}), 201


@app.route('/api/riego/<int:rid>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def manage_riego_one(rid):
    uid = get_uid()
    conn = get_db()
    if request.method == 'DELETE':
        conn.execute(
            "UPDATE riego SET deleted_at=datetime('now') WHERE id=? AND user_id=?",
            (rid, uid))
        conn.commit(); conn.close(); return jsonify({"status": "ok"})
    if request.method == 'GET':
        row = one(conn, "SELECT * FROM riego WHERE id=? AND user_id=? AND deleted_at IS NULL", (rid, uid))
        conn.close(); return jsonify(row or {})
    data = request.json or {}
    err = _validate_riego(data)
    if err:
        conn.close()
        return jsonify({"error": err}), 400
    fields = ['parcela_id','parcela_etiqueta','fecha','tipo_riego','volumen_m3',
              'horas_riego','fuente_agua','notas','campana']
    sets = ', '.join(f"{f}=?" for f in fields)
    numeric = {'volumen_m3', 'horas_riego'}
    values = [data.get(f) or None if f in numeric else data.get(f) for f in fields]
    conn.execute(f"UPDATE riego SET {sets} WHERE id=? AND user_id=? AND deleted_at IS NULL",
                 values + [rid, uid])
    conn.commit(); conn.close(); return jsonify({"status": "ok"})


# ─────────────────────────────────────────────
# PLAN DE ABONADO
# ─────────────────────────────────────────────
@app.route('/api/abonado', methods=['GET', 'POST'])
@login_required
def manage_abonado():
    uid = get_uid()
    conn = get_db()
    if request.method == 'GET':
        campana = request.args.get('campana', '2025/2026')
        rows = dicts(conn, "SELECT * FROM abonado WHERE user_id=? AND campana=? AND deleted_at IS NULL ORDER BY fecha_preparacion DESC", (uid, campana))
        conn.close(); return jsonify(rows)

    data = request.json or {}
    err = _validate_abonado(data)
    if err:
        conn.close()
        return jsonify({"error": err}), 400

    c = conn.cursor()
    c.execute('''
        INSERT INTO abonado (
            user_id, parcela_id, parcela_etiqueta, cultivo, cultivo_anterior,
            rendimiento_esperado_kg_ha, n_necesario_kg_ha, p_necesario_kg_ha,
            k_necesario_kg_ha, fecha_preparacion, datos_suelo,
            abono_recomendado, dosis_recomendada_kg_ha, notas, campana
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    ''', (uid, data.get('parcela_id'), data.get('parcela_etiqueta'),
          data.get('cultivo'), data.get('cultivo_anterior'),
          data.get('rendimiento_esperado_kg_ha') or None,
          data.get('n_necesario_kg_ha'), data.get('p_necesario_kg_ha'), data.get('k_necesario_kg_ha'),
          data.get('fecha_preparacion'), data.get('datos_suelo'),
          data.get('abono_recomendado'), data.get('dosis_recomendada_kg_ha') or None,
          data.get('notas'), data.get('campana', '2025/2026')))
    conn.commit(); new_id = c.lastrowid; conn.close()
    return jsonify({"status": "ok", "id": new_id}), 201


@app.route('/api/abonado/<int:aid>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def manage_abonado_one(aid):
    uid = get_uid()
    conn = get_db()
    if request.method == 'DELETE':
        conn.execute(
            "UPDATE abonado SET deleted_at=datetime('now') WHERE id=? AND user_id=?",
            (aid, uid))
        conn.commit(); conn.close(); return jsonify({"status": "ok"})
    if request.method == 'GET':
        row = one(conn, "SELECT * FROM abonado WHERE id=? AND user_id=? AND deleted_at IS NULL", (aid, uid))
        conn.close(); return jsonify(row or {})
    data = request.json or {}
    err = _validate_abonado(data)
    if err:
        conn.close()
        return jsonify({"error": err}), 400
    fields = ['parcela_id', 'parcela_etiqueta', 'cultivo', 'cultivo_anterior',
              'rendimiento_esperado_kg_ha', 'n_necesario_kg_ha', 'p_necesario_kg_ha',
              'k_necesario_kg_ha', 'fecha_preparacion', 'datos_suelo',
              'abono_recomendado', 'dosis_recomendada_kg_ha', 'notas', 'campana']
    numeric = {'rendimiento_esperado_kg_ha', 'n_necesario_kg_ha', 'p_necesario_kg_ha',
               'k_necesario_kg_ha', 'dosis_recomendada_kg_ha'}
    sets = ', '.join(f"{f}=?" for f in fields)
    values = [data.get(f) or None if f in numeric else data.get(f) for f in fields]
    conn.execute(f"UPDATE abonado SET {sets} WHERE id=? AND user_id=? AND deleted_at IS NULL",
                 values + [aid, uid])
    conn.commit(); conn.close(); return jsonify({"status": "ok"})


# ─────────────────────────────────────────────
# LABORES
# ─────────────────────────────────────────────
@app.route('/api/labores', methods=['GET', 'POST'])
@login_required
def manage_labores():
    uid = get_uid()
    conn = get_db()
    if request.method == 'GET':
        rows = dicts(conn, "SELECT * FROM labores WHERE user_id=? ORDER BY fecha DESC", (uid,))
        conn.close(); return jsonify(rows)
    data = request.json or {}
    c = conn.cursor()
    c.execute('''
        INSERT INTO labores (user_id, parcela_id, parcela_etiqueta, fecha,
            tipo_labor, descripcion, producto, maquinaria, horas_trabajadas, operario, notas, campana)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
    ''', (uid, data.get('parcela_id'), data.get('parcela_etiqueta'), data.get('fecha'),
          data.get('tipo_labor'), data.get('descripcion'), data.get('producto'), data.get('maquinaria'),
          data.get('horas_trabajadas') or None, data.get('operario'), data.get('notas'),
          data.get('campana', '2025/2026')))
    conn.commit(); new_id = c.lastrowid; conn.close()
    return jsonify({"status": "ok", "id": new_id}), 201


@app.route('/api/labores/<int:lid>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def manage_labor(lid):
    uid = get_uid()
    conn = get_db()
    if request.method == 'DELETE':
        conn.execute("DELETE FROM labores WHERE id=? AND user_id=?", (lid, uid))
        conn.commit(); conn.close(); return jsonify({"status": "ok"})
    if request.method == 'GET':
        row = one(conn, "SELECT * FROM labores WHERE id=? AND user_id=?", (lid, uid))
        conn.close(); return jsonify(row or {})
    data = request.json or {}
    fields = ['parcela_id','parcela_etiqueta','fecha','tipo_labor','descripcion',
              'producto','maquinaria','horas_trabajadas','operario','notas','campana']
    sets = ', '.join(f"{f}=?" for f in fields)
    conn.execute(f"UPDATE labores SET {sets} WHERE id=? AND user_id=?",
                 [data.get(f) or None if f == 'horas_trabajadas' else data.get(f) for f in fields] + [lid, uid])
    conn.commit(); conn.close(); return jsonify({"status": "ok"})


# ─────────────────────────────────────────────
# COSECHA
# ─────────────────────────────────────────────
@app.route('/api/cosecha', methods=['GET', 'POST'])
@login_required
def manage_cosecha():
    uid = get_uid()
    conn = get_db()
    if request.method == 'GET':
        rows = dicts(conn, "SELECT * FROM cosecha WHERE user_id=? ORDER BY fecha_inicio DESC", (uid,))
        conn.close(); return jsonify(rows)
    data = request.json or {}

    # Validar fechas de cosecha
    fi = data.get('fecha_inicio')
    ff = data.get('fecha_fin')
    if not fi:
        conn.close()
        return jsonify({"error": "La fecha de inicio es obligatoria"}), 400
    try:
        d_ini = datetime.date.fromisoformat(str(fi))
        if d_ini > datetime.date.today():
            conn.close()
            return jsonify({"error": "La fecha de inicio de cosecha no puede ser futura"}), 400
        if ff:
            d_fin = datetime.date.fromisoformat(str(ff))
            if d_fin > datetime.date.today():
                conn.close()
                return jsonify({"error": "La fecha de fin de cosecha no puede ser futura"}), 400
            if d_fin < d_ini:
                conn.close()
                return jsonify({"error": "La fecha de fin no puede ser anterior a la de inicio"}), 400
    except (ValueError, TypeError):
        conn.close()
        return jsonify({"error": "Formato de fecha inválido (use YYYY-MM-DD)"}), 400

    # Bloquear cosecha si hay tratamientos con plazo de seguridad no vencido en la misma parcela
    if data.get('parcela_id') and data.get('fecha_inicio'):
        plazo_activos = dicts(conn, """
            SELECT producto_comercial, fecha_recoleccion_minima
            FROM tratamientos
            WHERE parcela_id=? AND user_id=? AND deleted_at IS NULL
              AND fecha_recoleccion_minima > ?
        """, (data['parcela_id'], uid, data['fecha_inicio']))
        if plazo_activos:
            nombres = ', '.join(p['producto_comercial'] for p in plazo_activos)
            conn.close()
            return jsonify({"error": (
                f"No se puede registrar la cosecha: plazo de seguridad no vencido para: {nombres}. "
                "Espera hasta que pasen los plazos indicados en los tratamientos."
            )}), 400

    prod = data.get('produccion_total_valor') or 0
    sup  = data.get('superficie_cosechada_ha') or 0
    rend = round(float(prod) / float(sup), 2) if sup and float(sup) > 0 else None
    c = conn.cursor()
    c.execute('''
        INSERT INTO cosecha (user_id, parcela_id, parcela_etiqueta, fecha_inicio, fecha_fin,
            cultivo, variedad, superficie_cosechada_ha, produccion_total_valor,
            produccion_total_unidad, rendimiento_kg_ha, destino, comprador,
            precio_unidad, notas, campana)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    ''', (uid, data.get('parcela_id'), data.get('parcela_etiqueta'),
          data.get('fecha_inicio'), data.get('fecha_fin'), data.get('cultivo'),
          data.get('variedad'), sup, prod, data.get('produccion_total_unidad', 'kg'),
          rend, data.get('destino'), data.get('comprador'),
          data.get('precio_unidad') or None, data.get('notas'), data.get('campana', '2025/2026')))
    conn.commit(); new_id = c.lastrowid; conn.close()
    return jsonify({"status": "ok", "id": new_id}), 201


@app.route('/api/cosecha/<int:cid>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def manage_cosecha_one(cid):
    uid = get_uid()
    conn = get_db()
    if request.method == 'DELETE':
        conn.execute("DELETE FROM cosecha WHERE id=? AND user_id=?", (cid, uid))
        conn.commit(); conn.close(); return jsonify({"status": "ok"})
    if request.method == 'GET':
        row = one(conn, "SELECT * FROM cosecha WHERE id=? AND user_id=?", (cid, uid))
        conn.close(); return jsonify(row or {})
    data = request.json or {}
    fields = ['parcela_id','parcela_etiqueta','fecha_inicio','fecha_fin','cultivo',
              'variedad','superficie_cosechada_ha','produccion_total_valor','produccion_total_unidad',
              'rendimiento_kg_ha','destino','comprador','precio_unidad','notas','campana']
    _numeric_c = {'superficie_cosechada_ha', 'produccion_total_valor', 'rendimiento_kg_ha', 'precio_unidad'}
    sets = ', '.join(f"{f}=?" for f in fields)
    conn.execute(f"UPDATE cosecha SET {sets} WHERE id=? AND user_id=?",
                 [data.get(f) or None if f in _numeric_c else data.get(f) for f in fields] + [cid, uid])
    conn.commit(); conn.close(); return jsonify({"status": "ok"})


# ─────────────────────────────────────────────
# EQUIPOS
# ─────────────────────────────────────────────
@app.route('/api/equipos', methods=['GET', 'POST'])
@login_required
def manage_equipos():
    uid = get_uid()
    conn = get_db()
    if request.method == 'GET':
        rows = dicts(conn, "SELECT * FROM equipos WHERE user_id=?", (uid,))
        conn.close(); return jsonify(rows)
    data = request.json or {}
    c = conn.cursor()
    c.execute('''INSERT INTO equipos (user_id, descripcion, tipo, marca, modelo, num_registro_roma, fecha_iteaf, notas)
                 VALUES (?,?,?,?,?,?,?,?)''',
              (uid, data.get('descripcion'), data.get('tipo'), data.get('marca'),
               data.get('modelo'), data.get('num_registro_roma'), data.get('fecha_iteaf'), data.get('notas')))
    conn.commit(); new_id = c.lastrowid; conn.close()
    return jsonify({"status": "ok", "id": new_id}), 201


@app.route('/api/equipos/<int:eid>', methods=['PUT', 'DELETE'])
@login_required
def manage_equipo(eid):
    uid = get_uid()
    conn = get_db()
    if request.method == 'DELETE':
        conn.execute("DELETE FROM equipos WHERE id=? AND user_id=?", (eid, uid))
        conn.commit(); conn.close(); return jsonify({"status": "ok"})
    data = request.json or {}
    fields = ['descripcion','tipo','marca','modelo','num_registro_roma','fecha_iteaf','notas']
    sets = ', '.join(f"{f}=?" for f in fields)
    conn.execute(f"UPDATE equipos SET {sets} WHERE id=? AND user_id=?",
                 [data.get(f) for f in fields] + [eid, uid])
    conn.commit(); conn.close(); return jsonify({"status": "ok"})


# ─────────────────────────────────────────────
# APLICADORES
# ─────────────────────────────────────────────
@app.route('/api/aplicadores', methods=['GET', 'POST'])
@login_required
def manage_aplicadores():
    uid = get_uid()
    conn = get_db()
    if request.method == 'GET':
        rows = dicts(conn, "SELECT * FROM aplicadores WHERE user_id=? AND activo=1", (uid,))
        conn.close(); return jsonify(rows)
    data = request.json or {}
    c = conn.cursor()
    c.execute("INSERT INTO aplicadores (user_id, nombre, nif, num_ropo) VALUES (?,?,?,?)",
              (uid, data.get('nombre'), data.get('nif'), data.get('num_ropo')))
    conn.commit(); new_id = c.lastrowid; conn.close()
    return jsonify({"status": "ok", "id": new_id}), 201


@app.route('/api/aplicadores/<int:aid>', methods=['PUT', 'DELETE'])
@login_required
def manage_aplicador(aid):
    uid = get_uid()
    conn = get_db()
    if request.method == 'DELETE':
        conn.execute("UPDATE aplicadores SET activo=0 WHERE id=? AND user_id=?", (aid, uid))
        conn.commit(); conn.close(); return jsonify({"status": "ok"})
    data = request.json or {}
    conn.execute("UPDATE aplicadores SET nombre=?, nif=?, num_ropo=? WHERE id=? AND user_id=?",
                 (data.get('nombre'), data.get('nif'), data.get('num_ropo'), aid, uid))
    conn.commit(); conn.close(); return jsonify({"status": "ok"})


# ─────────────────────────────────────────────
# COMPRAS (Trazabilidad — Anexo III S5)
# ─────────────────────────────────────────────
@app.route('/api/compras', methods=['GET', 'POST'])
@login_required
def manage_compras():
    uid = get_uid()
    conn = get_db()
    if request.method == 'GET':
        campana = request.args.get('campana', '')
        sql = "SELECT * FROM compras WHERE user_id=? AND deleted_at IS NULL"
        params = [uid]
        if campana:
            sql += " AND campana=?"
            params.append(campana)
        sql += " ORDER BY fecha DESC"
        rows = dicts(conn, sql, params)
        conn.close(); return jsonify(rows)

    data = request.json or {}
    err = _validate_compra(data) or _validate_campana(data.get('campana'))
    if err:
        conn.close()
        return jsonify({"error": err}), 400

    c = conn.cursor()
    c.execute('''
        INSERT INTO compras (user_id, fecha, tipo_producto, producto, num_registro_mapa,
            sustancia_activa, proveedor, cantidad_valor, cantidad_unidad, num_lote,
            num_factura, precio_total, campana, notas)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    ''', (uid, data.get('fecha'), data.get('tipo_producto'), data.get('producto'),
          data.get('num_registro_mapa'), data.get('sustancia_activa'),
          data.get('proveedor'), data.get('cantidad_valor') or None,
          data.get('cantidad_unidad', 'kg'), data.get('num_lote'),
          data.get('num_factura'), data.get('precio_total') or None,
          data.get('campana', '2025/2026'), data.get('notas')))
    conn.commit(); new_id = c.lastrowid; conn.close()
    return jsonify({"status": "ok", "id": new_id}), 201


@app.route('/api/compras/<int:cid>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def manage_compra(cid):
    uid = get_uid()
    conn = get_db()
    if request.method == 'DELETE':
        conn.execute(
            "UPDATE compras SET deleted_at=datetime('now') WHERE id=? AND user_id=?",
            (cid, uid))
        conn.commit(); conn.close(); return jsonify({"status": "ok"})
    if request.method == 'GET':
        row = one(conn, "SELECT * FROM compras WHERE id=? AND user_id=? AND deleted_at IS NULL", (cid, uid))
        conn.close(); return jsonify(row or {})
    data = request.json or {}
    err = _validate_compra(data)
    if err:
        conn.close()
        return jsonify({"error": err}), 400
    fields = ['fecha','tipo_producto','producto','num_registro_mapa','sustancia_activa',
              'proveedor','cantidad_valor','cantidad_unidad','num_lote','num_factura',
              'precio_total','campana','notas']
    sets = ', '.join(f"{f}=?" for f in fields)
    _numeric_co = {'cantidad_valor', 'precio_total'}
    conn.execute(f"UPDATE compras SET {sets} WHERE id=? AND user_id=? AND deleted_at IS NULL",
                 [data.get(f) or None if f in _numeric_co else data.get(f) for f in fields] + [cid, uid])
    conn.commit(); conn.close(); return jsonify({"status": "ok"})


# ─────────────────────────────────────────────
# SIGPAC PROXY
# ─────────────────────────────────────────────
SIGPAC_BASE = "https://sigpac.mapa.gob.es/fega/serviciosvisorsigpac/query"

def _sigpac_get(url, timeout=10):
    try:
        r = req_lib.get(url, timeout=timeout)
        r.raise_for_status()
        return r.json()
    except Exception as e:
        return {"error": str(e)}

def _sigpac_list(url):
    import re
    data = _sigpac_get(url)
    if not isinstance(data, dict) or 'features' not in data:
        return data
    out = []
    for f in data.get('features', []):
        p = (f or {}).get('properties') or {}
        nombre = p.get('nombre') or ''
        nombre = re.sub(r'\s*\(\d+\)\s*$', '', nombre).strip()
        out.append({'codigo': p.get('codigo'), 'nombre': nombre})
    out.sort(key=lambda x: (x['nombre'] or '').lower())
    return out

def _sigpac_param(val, default=''):
    """Valida que un parámetro SIGPAC sea solo dígitos (máx 6). Evita inyección en URLs."""
    s = str(val).strip() if val else default
    if not re.fullmatch(r'\d{1,6}', s):
        return None
    return s

@app.route('/api/sigpac/provincias')
@login_required
def sigpac_provincias():
    return jsonify(_sigpac_get(f"{SIGPAC_BASE}/provincias"))

@app.route('/api/sigpac/municipios')
@login_required
def sigpac_municipios():
    prov = _sigpac_param(request.args.get('provincia_cod'), '13')
    if not prov: return jsonify({"error": "Parámetro provincia inválido"}), 400
    return jsonify(_sigpac_list(f"{SIGPAC_BASE}/municipios/{prov}"))

@app.route('/api/sigpac/poligonos')
@login_required
def sigpac_poligonos():
    prov = _sigpac_param(request.args.get('provincia_cod'), '13')
    mun  = _sigpac_param(request.args.get('municipio_cod'), '131')
    if not prov or not mun: return jsonify({"error": "Parámetros SIGPAC inválidos"}), 400
    return jsonify(_sigpac_get(f"{SIGPAC_BASE}/poligonos/{prov}/{mun}"))

@app.route('/api/sigpac/parcelas')
@login_required
def sigpac_parcelas():
    prov = _sigpac_param(request.args.get('provincia_cod'), '13')
    mun  = _sigpac_param(request.args.get('municipio_cod'), '131')
    pol  = _sigpac_param(request.args.get('poligono'), '1')
    if not prov or not mun or not pol: return jsonify({"error": "Parámetros SIGPAC inválidos"}), 400
    return jsonify(_sigpac_get(f"{SIGPAC_BASE}/parcelas/{prov}/{mun}/{pol}"))

@app.route('/api/sigpac/recintos')
@login_required
def sigpac_recintos():
    prov = _sigpac_param(request.args.get('provincia'), '13')
    mun  = _sigpac_param(request.args.get('municipio'), '131')
    pol  = _sigpac_param(request.args.get('poligono'), '1')
    par  = _sigpac_param(request.args.get('parcela'), '1')
    agr  = _sigpac_param(request.args.get('agregado'), '0')
    zona = _sigpac_param(request.args.get('zona'), '0')
    if not all([prov, mun, pol, par, agr, zona]): return jsonify({"error": "Parámetros SIGPAC inválidos"}), 400
    return jsonify(_sigpac_get(f"{SIGPAC_BASE}/recintos/{prov}/{mun}/{agr}/{zona}/{pol}/{par}"))


# Mapa de código cultivo Catastro → uso SIGPAC
_CATASTRO_A_SIGPAC = {
    'O-': 'OV-OLIVAR',  'OL': 'OV-OLIVAR',
    'VI': 'VI-VIÑEDO',  'V-': 'VI-VIÑEDO',
    'TA': 'TA-TIERRA ARABLE', 'TH': 'TA-TIERRA ARABLE',
    'CF': 'CF-CITRICOS', 'CI': 'CI-CITRICOS-INVER',
    'CS': 'CS-CULTIVOS SIN ESPECIF',
    'FF': 'FL-FRUTOS SECOS', 'FL': 'FL-FRUTOS SECOS',
    'FY': 'FY-FRUTALES',
    'PA': 'PA-PASTO', 'PR': 'PR-PASTO ARBUSTIVO', 'PS': 'PS-PASTIZAL',
    'CA': 'CA-VIALES', 'IM': 'IM-IMPRODUCTIVO',
    'ZU': 'ZU-ZONA URBANA', 'AG': 'AG-CORRIENTE AGUA',
}

def _catastro_a_uso_sigpac(ccc, dcc=''):
    """Convierte código cultivo Catastro a uso SIGPAC."""
    if ccc:
        for k, v in _CATASTRO_A_SIGPAC.items():
            if ccc.upper().startswith(k):
                return v
    # Fallback por descripción
    dcc_n = (dcc or '').upper()
    if 'OLIVO' in dcc_n: return 'OV-OLIVAR'
    if 'VIÑA' in dcc_n or 'VID' in dcc_n: return 'VI-VIÑEDO'
    if 'CEREAL' in dcc_n or 'TRIGO' in dcc_n or 'CEBADA' in dcc_n: return 'TA-TIERRA ARABLE'
    if 'ALMENDRO' in dcc_n or 'FRUTO SECO' in dcc_n: return 'FL-FRUTOS SECOS'
    if 'CITRICO' in dcc_n or 'NARANJO' in dcc_n: return 'CF-CITRICOS'
    if 'PASTIZAL' in dcc_n or 'PASTO' in dcc_n: return 'PA-PASTO'
    return ''

_USO_LABELS = {
    'OV':'OV-OLIVAR','VI':'VI-VIÑEDO','TA':'TA-TIERRA ARABLE',
    'TH':'TH-HUERTA','CF':'CF-CITRICOS','FL':'FL-FRUTOS SECOS',
    'FY':'FY-FRUTALES','PA':'PA-PASTO','PR':'PR-PASTO ARBUSTIVO',
    'PS':'PS-PASTIZAL','CA':'CA-VIALES','IM':'IM-IMPRODUCTIVO',
    'AG':'AG-CORRIENTES AGUA','ZU':'ZU-ZONA URBANA','ED':'ED-EDIFICACIONES',
    'IV':'IV-INVERNADERO',
}

@app.route('/api/sigpac/datos')
@login_required
def sigpac_datos():
    """Obtiene superficie y uso SIGPAC via endpoint de intersección."""
    prov = _sigpac_param(request.args.get('provincia'), '')
    mun  = _sigpac_param(request.args.get('municipio'), '')
    pol  = _sigpac_param(request.args.get('poligono'), '')
    par  = _sigpac_param(request.args.get('parcela'), '')
    rec  = _sigpac_param(request.args.get('recinto') or '1', '1')
    if not all([prov, mun, pol, par, rec]):
        return jsonify({"error": "Parámetros SIGPAC inválidos"}), 400

    resultado = {'superficie_ha': '', 'uso_sigpac': '', 'referencia_cat': '', 'num_recintos': 0}

    try:
        # Paso 1: contar recintos
        recintos_data = _sigpac_get(f"{SIGPAC_BASE}/recintos/{prov}/{mun}/0/0/{pol}/{par}")
        resultado['num_recintos'] = len(recintos_data.get('features', []))

        # Paso 2: detalle del recinto por referencia completa
        # Este es el único endpoint que funciona y devuelve parcelaInfo
        INTER = "https://sigpac.mapa.gob.es/fega/serviciosvisorsigpac/intersection"
        inter = _sigpac_get(f"{INTER}/recinto/recinto/{prov},{mun},0,0,{pol},{par},{rec}")
        pi = inter.get('parcelaInfo') or {}

        dn = pi.get('dn_surface')
        if dn:
            resultado['superficie_ha'] = round(float(dn) / 10000, 4)

        ref_cat = pi.get('referencia_cat', '')
        resultado['referencia_cat'] = ref_cat

        # Paso 3: uso SIGPAC via Catastro (parcelaInfo no lo incluye)
        if ref_cat:
            try:
                import xml.etree.ElementTree as ET
                cat_r = req_lib.get(
                    'https://ovc.catastro.meh.es/ovcservweb/OVCSWLocalizacionRC/OVCCallejero.asmx/Consulta_DNPRC',
                    params={'Provincia': '', 'Municipio': '', 'RC': ref_cat},
                    timeout=8
                )
                ns = {'c': 'http://www.catastro.meh.es/'}
                root = ET.fromstring(cat_r.text)
                ccc = root.find('.//c:ccc', ns)
                dcc = root.find('.//c:dcc', ns)
                uso = _catastro_a_uso_sigpac(
                    ccc.text if ccc is not None else '',
                    dcc.text if dcc is not None else ''
                )
                if uso:
                    resultado['uso_sigpac'] = uso
                    resultado['cultivo_catastro'] = dcc.text if dcc is not None else ''
            except Exception:
                pass  # Catastro es opcional

    except Exception as e:
        resultado['error'] = str(e)

    return jsonify(resultado)


@app.route('/api/sigpac/debug')
@login_required
@admin_required
def sigpac_debug():
    """Prueba todas las variantes de la API SIGPAC y devuelve respuestas crudas."""
    prov = request.args.get('provincia', '13')
    mun  = request.args.get('municipio', '')
    pol  = request.args.get('poligono', '')
    par  = request.args.get('parcela', '')
    rec  = request.args.get('recinto', '1')
    INTER = "https://sigpac.mapa.gob.es/fega/serviciosvisorsigpac/intersection"

    recintos_raw = _sigpac_get(f"{SIGPAC_BASE}/recintos/{prov}/{mun}/0/0/{pol}/{par}")
    features = recintos_raw.get('features', [])
    dn_pk = features[0].get('properties', {}).get('dn_pk') if features else None

    dk = str(dn_pk) if dn_pk else None
    return jsonify({
        'params': {'prov': prov, 'mun': mun, 'pol': pol, 'par': par, 'rec': rec},
        'recintos_features_count': len(features),
        'recintos_first_props': features[0].get('properties', {}) if features else {},
        'dn_pk': dk,
        'query_recinto_dnpk':    _sigpac_get(f"{SIGPAC_BASE}/recinto/{dk}") if dk else 'no dn_pk',
        'inter_recinto_dnpk':    _sigpac_get(f"{INTER}/recinto/{dk}") if dk else 'no dn_pk',
        'inter_geometria_dnpk':  _sigpac_get(f"{INTER}/recinto/geometria/{dk}") if dk else 'no dn_pk',
        'inter_by_ref':          _sigpac_get(f"{INTER}/recinto/recinto/{prov},{mun},0,0,{pol},{par},{rec}"),
    })


# ─────────────────────────────────────────────
# NLP SIMPLE — PARSING DE TEXTO LIBRE
# ─────────────────────────────────────────────
import unicodedata as _UD
import re as _re

def _norm(s):
    """Minúsculas sin acentos para comparación robusta."""
    return ''.join(
        c for c in _UD.normalize('NFD', str(s).lower())
        if _UD.category(c) != 'Mn'
    )

def extraer_parcela(texto, uid):
    conn = get_db()
    parcelas = dicts(conn, "SELECT id, nombre_finca FROM parcelas WHERE user_id=? AND activa=1", (uid,))
    conn.close()
    tnorm = _norm(texto)

    # 1) Coincidencia exacta normalizada (sin acentos ni mayúsculas)
    for p in parcelas:
        if _norm(p['nombre_finca']) in tnorm:
            return {'id': p['id'], 'nombre': p['nombre_finca']}

    # 2) Coincidencia parcial: todas las palabras >3 letras del nombre están en el texto
    for p in parcelas:
        partes = [w for w in _re.split(r'[\s\-/]+', _norm(p['nombre_finca'])) if len(w) > 3]
        if partes and all(parte in tnorm for parte in partes):
            return {'id': p['id'], 'nombre': p['nombre_finca']}

    return None

def extraer_nombre_candidato(texto):
    """Extrae el candidato a nombre de parcela.
    Prioridad 1: locativo explícito 'en/finca/parcela/campo [el/la] NOMBRE'
    Prioridad 2: patrón verbo → NOMBRE (sin 'en' intermedio)
    """
    STOP_FIN = {'hoy', 'ayer', 'esta', 'este', 'manana', 'por', 'he', 'con', 'sin', 'y'}
    # Palabras función que se eliminan al inicio SOLO si están en minúscula.
    # Mayúscula inicial = parte del nombre propio (ej: "Las Mesas", "El Llano").
    SKIP_LOW = {'el', 'la', 'los', 'las', 'un', 'una', 'unos', 'unas',
                'parcela', 'finca', 'campo', 'terreno'}

    def _limpiar(palabras):
        while palabras and _norm(palabras[-1]) in {_norm(s) for s in STOP_FIN}:
            palabras.pop()
        while palabras and palabras[0].lower() in SKIP_LOW and not palabras[0][0].isupper():
            palabras.pop(0)
        return ' '.join(palabras)

    # ── Prioridad 1: locativo explícito ──
    # No consume el artículo en el regex; _limpiar decide si eliminarlo o no
    # según mayúscula/minúscula. Así "las Mesas"→"MESAS" y "Las Mesas"→"LAS MESAS".
    m1 = _re.search(
        r'(?:en|finca|parcela|campo)\s+([\w][\w\s]{1,30}?)'
        r'(?=\s+con\s|\s*[,.]|\s*$)',
        texto, _re.IGNORECASE
    )
    if m1:
        candidato = _limpiar(m1.group(1).strip().split())
        if 2 < len(candidato) < 40:
            return candidato.upper()

    # ── Prioridad 2: verbo → [artículo] → NOMBRE (sin preposición en medio) ──
    # Solo captura hasta 'en' para evitar arrastrar el producto.
    VERBOS = (
        r'trat\w{1,6}|abon\w{1,6}|reg\w{1,6}|pod\w{1,6}|cosech\w{1,6}|'
        r'fumig\w{1,6}|sembr\w{1,6}|siembr\w{1,5}|labr\w{1,6}|vendimi\w{1,6}|'
        r'recog\w{1,6}|arad\w{1,5}|cav\w{1,5}|desbroz\w{1,5}|pulver\w{1,5}|'
        r'trilla\w{1,4}|recolect\w{1,5}|subsolad\w{1,4}|grad\w{1,5}|cultiv\w{1,5}'
    )
    m2 = _re.search(
        r'(?:' + VERBOS + r')'
        r'\s+(?:el\s+|la\s+|los\s+|las\s+)?'
        r'([\w][\w\s]{1,33}?)'
        r'(?=\s+en\s|\s+con\s|\s*[,.]|\s+(?:hoy|ayer|esta|este|por|de\s+|$)|\s*$)',
        texto, _re.IGNORECASE
    )
    if m2:
        candidato = _limpiar(m2.group(1).strip().split())
        if 2 < len(candidato) < 40:
            return candidato.upper()

    return None

def extraer_accion(texto):
    tnorm = _norm(texto)
    acciones = {
        'tratamiento': [
            'tratado', 'tratamiento', 'tratamos', 'trate', 'pulverizado', 'pulverice',
            'fumigado', 'fumigue', 'fumigaci', 'spray', 'insecticida', 'fungicida',
            'herbicida', 'fitosanitario', 'plaguicida', 'mata', 'mato',
        ],
        'fertilizacion': [
            'abonado', 'abone', 'abonamos', 'fertilizado', 'fertilice', 'abono',
            'nutrientes', 'nitrogeno', 'fosforo', 'potasio', 'npk', 'urea',
            'sulfato', 'superfosfato', 'estiercol', 'compost', 'purines',
        ],
        'riego': [
            'regado', 'regue', 'rege', 'regamos', 'rego', 'riego', 'riegos',
            'inundado', 'goteo', 'aspersion', 'pivote',
        ],
        'cosecha': [
            'cosechado', 'cosechamos', 'coseche', 'cosecho', 'cosecha ',
            'recolectado', 'recogie', 'vendimiado', 'vendimia',
            'trillado', 'trilla', 'recogi',
        ],
        'labor': [
            'labor', 'labrado', 'labre', 'laboreo',
            'arado', 'are ', 'are,', 'cave', 'cavado',
            'poda', 'pode', 'podamos',
            'desyerbado', 'desherbado', 'desbroz',
            'sembrado', 'siembre', 'siembra', 'sembre', 'sembré', 'he sembrado',
            'he sembrado', 'sembramos', 'sembrar', 'siembro',
            'fresado', 'subsolado', 'cultivado', 'he cultivado', 'cultivé', 'cultivamos',
            'gradeo', 'pase de', 'pase ', 'limpie', 'limpieza',
            'plantado', 'plante', 'planté', 'plantamos', 'plantación',
        ],
    }
    for tipo, palabras in acciones.items():
        for palabra in palabras:
            if _norm(palabra) in tnorm:
                return {'tipo': tipo, 'confianza': 0.9, 'palabra_clave': palabra}
    return {'tipo': None, 'confianza': 0, 'palabra_clave': None}

def extraer_producto(texto):
    tnorm = _norm(texto)
    # (palabra_clave_normalizada, nombre_display)
    productos = [
        # Semillas / cultivos
        ('yero', 'Yeros'), ('yeros', 'Yeros'),
        ('trigo', 'Trigo'), ('cebada', 'Cebada'), ('avena', 'Avena'),
        ('centeno', 'Centeno'), ('triticale', 'Triticale'),
        ('girasol', 'Girasol'), ('colza', 'Colza'), ('maiz', 'Maíz'),
        ('soja', 'Soja'), ('guisante', 'Guisante'), ('garbanzo', 'Garbanzo'),
        ('lenteja', 'Lenteja'), ('almorta', 'Almorta'), ('veza', 'Veza'),
        ('alfalfa', 'Alfalfa'), ('remolacha', 'Remolacha'), ('patata', 'Patata'),
        ('tomate', 'Tomate'), ('pimiento', 'Pimiento'), ('cebolla', 'Cebolla'),
        ('ajo', 'Ajo'), ('olivo', 'Olivo'), ('vid', 'Vid'), ('viña', 'Vid'),
        ('almendro', 'Almendro'), ('pistachero', 'Pistachero'),
        # Fungicidas
        ('cobre', 'Cobre'), ('azufre', 'Azufre'), ('mancozeb', 'Mancozeb'),
        ('captan', 'Captán'), ('clorotalonil', 'Clorotalonil'), ('tebuconazol', 'Tebuconazol'),
        ('iprodiona', 'Iprodiona'), ('metalaxil', 'Metalaxil'), ('fosetil', 'Fosetil-Al'),
        ('ziram', 'Ziram'), ('metiram', 'Metiram'), ('oxicloruro', 'Oxicloruro de cobre'),
        # Insecticidas
        ('clorpirifos', 'Clorpirifos'), ('deltametrina', 'Deltametrina'),
        ('lambda', 'Lambda-cihalotrin'), ('imidacloprid', 'Imidacloprid'),
        ('spinosad', 'Spinosad'), ('abamectina', 'Abamectina'),
        ('dimetoato', 'Dimetoato'), ('piretrinas', 'Piretrinas'),
        # Herbicidas
        ('glifosato', 'Glifosato'), ('diquat', 'Diquat'), ('terbutilazina', 'Terbutilazina'),
        ('s-metolacloro', 'S-metolacloro'), ('pendimetalina', 'Pendimetalina'),
        # Fertilizantes
        ('urea', 'Urea'), ('npk', 'NPK'), ('nitrato amonico', 'Nitrato amónico'),
        ('sulfato amonico', 'Sulfato amónico'), ('superfosfato', 'Superfosfato'),
        ('cloruro potasico', 'Cloruro potásico'), ('estiercol', 'Estiércol'),
        ('compost', 'Compost'), ('purines', 'Purines'),
    ]
    for clave, nombre in productos:
        if _norm(clave) in tnorm:
            return {'nombre': nombre, 'confianza': 0.85}
    return {'nombre': None, 'confianza': 0}

def extraer_dosis(texto):
    patrones = [
        (r'(\d+[.,]?\d*)\s*(cc|centilitro)', 'cc'),
        (r'(\d+[.,]?\d*)\s*(l|litro|litros|ℓ)\b', 'L'),
        (r'(\d+[.,]?\d*)\s*(kg|kilo|kilos)\b', 'kg'),
        (r'(\d+[.,]?\d*)\s*(g|gramo|gramos)\b', 'g'),
        (r'(\d+[.,]?\d*)\s*(t|tonelada|toneladas)\b', 't'),
    ]
    for patron, unidad in patrones:
        m = _re.search(patron, texto, _re.IGNORECASE)
        if m:
            valor = float(m.group(1).replace(',', '.'))
            return {'valor': valor, 'unidad': unidad, 'texto_original': m.group(0)}
    return {'valor': None, 'unidad': None, 'texto_original': None}

def extraer_fecha(texto):
    """Extrae fecha del texto en lenguaje natural. Devuelve ISO string o hoy."""
    import datetime
    texto_l = texto.lower()
    hoy = datetime.date.today()

    if 'anteayer' in texto_l:
        return (hoy - datetime.timedelta(days=2)).isoformat()
    if 'ayer' in texto_l:
        return (hoy - datetime.timedelta(days=1)).isoformat()

    MESES = {
        'enero':1,'febrero':2,'marzo':3,'abril':4,'mayo':5,'junio':6,
        'julio':7,'agosto':8,'septiembre':9,'octubre':10,'noviembre':11,'diciembre':12
    }
    m = _re.search(r'\b(\d{1,2})\s+de\s+(' + '|'.join(MESES.keys()) + r')\b', texto_l)
    if m:
        dia, mes = int(m.group(1)), MESES[m.group(2)]
        for yr in (hoy.year, hoy.year - 1):
            try:
                fecha = datetime.date(yr, mes, dia)
                if fecha <= hoy + datetime.timedelta(days=1):
                    return fecha.isoformat()
            except ValueError:
                pass

    m = _re.search(r'\b(\d{1,2})[/-](\d{1,2})\b', texto_l)
    if m:
        dia, mes = int(m.group(1)), int(m.group(2))
        for yr in (hoy.year, hoy.year - 1):
            try:
                fecha = datetime.date(yr, mes, dia)
                if fecha <= hoy + datetime.timedelta(days=1):
                    return fecha.isoformat()
            except ValueError:
                pass

    DIAS = {
        'lunes':0,'martes':1,'miércoles':2,'miercoles':2,
        'jueves':3,'viernes':4,'sábado':5,'sabado':5,'domingo':6
    }
    for nombre, num in DIAS.items():
        if nombre in texto_l:
            atras = (hoy.weekday() - num) % 7 or 7
            return (hoy - datetime.timedelta(days=atras)).isoformat()

    return hoy.isoformat()


@app.route('/api/parse', methods=['POST'])
@login_required
def parse_texto_libre():
    uid = get_uid()
    data = request.json or {}
    texto = (data.get('texto') or '').strip()

    if not texto:
        return jsonify({"status": "error", "message": "El texto no puede estar vacío"}), 400

    parcela_data = extraer_parcela(texto, uid)
    accion_data = extraer_accion(texto)
    producto_data = extraer_producto(texto)
    dosis_data = extraer_dosis(texto)
    nombre_candidato = None if parcela_data else extraer_nombre_candidato(texto)

    parcela_id = parcela_data['id'] if parcela_data else None
    fecha = extraer_fecha(texto)

    return jsonify({
        "status": "success",
        "texto_original": texto,
        "parseo": {
            "parcela": {
                "id": parcela_id,
                "nombre": parcela_data['nombre'] if parcela_data else None,
                "nombre_candidato": nombre_candidato,
                "es_nueva": not parcela_data and bool(nombre_candidato),
                "requiere_seleccion": not parcela_data and not nombre_candidato,
                "confianza": 1.0 if parcela_data else 0.0,
            },
            "accion": {"tipo": accion_data['tipo'], "palabra_clave": accion_data['palabra_clave'], "confianza": accion_data['confianza']},
            "producto": {"nombre": producto_data['nombre'], "confianza": producto_data['confianza']},
            "dosis": {"valor": dosis_data['valor'], "unidad": dosis_data['unidad']},
            "fecha": fecha,
        },
        "requiere_confirmacion": not parcela_data and not nombre_candidato,
    })

@app.route('/api/parse/guardar', methods=['POST'])
@login_required
def parse_guardar():
    uid = get_uid()
    data = request.json or {}

    accion        = data.get('accion')
    palabra_clave = (data.get('palabra_clave') or '').strip()
    parcela_id    = data.get('parcela_id')
    parcela_nombre = (data.get('parcela_nombre') or '').strip()
    producto     = (data.get('producto') or '').strip()
    fecha        = data.get('fecha') or datetime.date.today().isoformat()
    texto        = (data.get('texto_original') or '').strip()
    campana      = data.get('campana') or '2025/2026'

    conn = get_db()

    # Auto-crear parcela si no existe
    if not parcela_id and parcela_nombre:
        c = conn.cursor()
        c.execute("INSERT INTO parcelas (user_id, nombre_finca) VALUES (?, ?)", (uid, parcela_nombre))
        parcela_id = c.lastrowid
        conn.commit()

    if not parcela_id:
        conn.close()
        return jsonify({"ok": False, "error": "No se pudo determinar la parcela. Indícala manualmente."}), 400

    parcela = one(conn, "SELECT nombre_finca FROM parcelas WHERE id=? AND user_id=?", (parcela_id, uid))
    etiqueta = (parcela or {}).get('nombre_finca', '')

    nota = f"NLP: {texto}" if texto else ''

    if accion == 'tratamiento':
        # Los tratamientos fitosanitarios requieren campos obligatorios del RD 1311/2012 (ROPO,
        # nº MAPA, sustancia activa, dosis, plazo de seguridad) que el texto libre no puede capturar.
        # Redirigir al formulario completo para evitar registros ilegalmente incompletos.
        conn.close()
        return jsonify({
            "ok": False,
            "requiere_formulario": True,
            "error": (
                "Los tratamientos fitosanitarios requieren datos obligatorios (Nº MAPA, sustancia "
                "activa, dosis, ROPO del aplicador, plazo de seguridad) que no se pueden capturar "
                "de texto libre. Usa el formulario completo para registrar este tratamiento."
            ),
            "producto": producto,
            "parcela_id": parcela_id,
            "fecha": fecha,
        }), 422
    elif accion == 'fertilizacion':
        # Igual que tratamientos: tipo de fertilizante, dosis y parcela son obligatorios por ley.
        conn.close()
        return jsonify({
            "ok": False,
            "requiere_formulario": True,
            "error": (
                "Las fertilizaciones requieren datos obligatorios (tipo de fertilizante, dosis, "
                "parcela) según el RD 1311/2012 Anexo III S4. Usa el formulario completo."
            ),
            "producto": producto,
            "parcela_id": parcela_id,
            "fecha": fecha,
        }), 422
    else:
        # palabra_clave viene del frontend; si no, re-extraer del texto
        if not palabra_clave:
            accion_det = extraer_accion(texto)
            palabra_clave = accion_det.get('palabra_clave') or ''
        # Normalizar al valor exacto del desplegable del formulario
        _LABOR_MAP = {
            'arado':'Arado','are':'Arado','arar':'Arado',
            'poda':'Poda','pode':'Poda','podamos':'Poda',
            'desherbado':'Desherbado','desyerbado':'Desherbado','desbroz':'Desherbado',
            'siembra':'Siembra','sembrado':'Siembra','siembre':'Siembra','sembre':'Siembra',
            'sembré':'Siembra','sembramos':'Siembra','sembrar':'Siembra','siembro':'Siembra',
            'he sembrado':'Siembra','plantado':'Plantación','plante':'Plantación',
            'planté':'Plantación','plantamos':'Plantación','plantación':'Plantación',
            'cultivado':'Siembra','cultivé':'Siembra','cultivamos':'Siembra','he cultivado':'Siembra',
            'fresado':'Fresado','fresa':'Fresado',
            'subsolado':'Subsolado',
            'gradeo':'Gradeo','pase':'Gradeo',
            'limpieza':'Limpieza','limpie':'Limpieza',
            'laboreo':'Laboreo del suelo','labrado':'Laboreo del suelo',
            'labor':'Laboreo del suelo',
            'cultivado':'Siembra','cultivé':'Siembra','cultivamos':'Siembra','he cultivado':'Siembra',
            'sembré':'Siembra','sembramos':'Siembra','sembrar':'Siembra','siembro':'Siembra',
            'he sembrado':'Siembra','plantado':'Plantación','plante':'Plantación',
            'planté':'Plantación','plantamos':'Plantación','plantación':'Plantación',
            'vendimia':'Vendimia','vendimiado':'Vendimia',
            'escarda':'Escarda','cave':'Escarda','cavado':'Escarda',
            'trilla':'Triturado de restos','trillado':'Triturado de restos',
            'riego':'Riego','regado':'Riego',
            'siega':'Otros','segado':'Otros',
        }
        tipo = _LABOR_MAP.get((palabra_clave or '').lower().strip()) or palabra_clave or accion or 'Otros'
        # Si es siembra/plantación y se detectó cultivo, usarlo como descripción
        if tipo in ('Siembra', 'Plantación') and producto:
            descripcion_labor = f"{tipo} de {producto}"
        else:
            descripcion_labor = texto
        conn.execute(
            "INSERT INTO labores (user_id, parcela_id, parcela_etiqueta, fecha, tipo_labor, descripcion, producto, notas, campana) VALUES (?,?,?,?,?,?,?,?,?)",
            (uid, parcela_id, etiqueta, fecha, tipo, descripcion_labor, producto or None, nota, campana)
        )

    conn.commit()
    conn.close()
    return jsonify({"ok": True, "parcela_id": parcela_id, "parcela_nombre": etiqueta or parcela_nombre})


# ─────────────────────────────────────────────
# IMPORT EXCEL
# ─────────────────────────────────────────────
def _importar_parcelas_wb(wb, uid):
    """Importa parcelas desde un workbook openpyxl. Devuelve (n_importadas, n_errores, detalle_errores)."""
    import unicodedata
    def _norm(s):
        return unicodedata.normalize('NFKD', str(s or '').lower().strip()).encode('ascii','ignore').decode()

    # Buscar hoja de parcelas (acepta cualquier nombre que contenga "parcela")
    ws = None
    for name in wb.sheetnames:
        if 'parcela' in _norm(name) or name == wb.sheetnames[0]:
            ws = wb[name]; break
    if not ws:
        return 0, 0, []

    # Leer cabeceras (fila 1) y mapear columnas por nombre
    headers = [_norm(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    def _col(keywords):
        for kw in keywords:
            for i, h in enumerate(headers):
                if kw in h: return i
        return None

    ci_nombre    = _col(['nombre','finca','parcela nom'])
    ci_provincia = _col(['provincia','prov'])
    ci_municipio = _col(['municipio','muni'])
    ci_poligono  = _col(['poligon','polig'])
    ci_parcela   = _col(['num parcela','parcela','numero','parcela num'])
    ci_recinto   = _col(['recinto'])

    faltan = [n for n, c in [('Nombre',ci_nombre),('Provincia',ci_provincia),('Municipio',ci_municipio),('Polígono',ci_poligono),('Parcela',ci_parcela)] if c is None]
    if faltan:
        return 0, 0, [f"Faltan columnas: {', '.join(faltan)}"]

    # Cache de códigos provincia/municipio para no consultar SIGPAC en cada fila
    conn = get_db()
    cur = conn.cursor()
    _cache_prov = {}; _cache_mun = {}

    def _get_prov_cod(nombre_prov):
        k = _norm(nombre_prov)
        if k not in _cache_prov:
            data = _sigpac_get(f"{SIGPAC_BASE}/provincias", timeout=4)
            for f in data.get('features', []):
                p = f.get('properties', {})
                _cache_prov[_norm(p.get('nombre',''))] = str(p.get('codigo',''))
        return _cache_prov.get(k, '')

    def _get_mun_cod(prov_cod, nombre_mun):
        k = f"{prov_cod}:{_norm(nombre_mun)}"
        if k not in _cache_mun:
            data = _sigpac_get(f"{SIGPAC_BASE}/municipios/{prov_cod}", timeout=4)
            for f in data.get('features', []):
                p = f.get('properties', {})
                _cache_mun[f"{prov_cod}:{_norm(p.get('nombre',''))}"] = str(p.get('codigo',''))
        return _cache_mun.get(k, '')

    n_ok = 0; errores = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row): continue
        nombre    = str(row[ci_nombre]    or '').strip()
        prov_nom  = str(row[ci_provincia] or '').strip()
        mun_nom   = str(row[ci_municipio] or '').strip()
        poligono  = str(row[ci_poligono]  or '').strip()
        parcela   = str(row[ci_parcela]   or '').strip()
        recinto   = str(row[ci_recinto]   or '').strip() if ci_recinto is not None else '1'

        if not nombre or not poligono or not parcela or not prov_nom or not mun_nom:
            errores.append(f"Fila incompleta: '{nombre or '?'}'")
            continue

        prov_cod = _get_prov_cod(prov_nom)
        mun_cod  = _get_mun_cod(prov_cod, mun_nom) if prov_cod else ''

        # Lookup SIGPAC para superficie y uso
        sup_ha = None; uso_sigpac = ''
        if prov_cod and mun_cod:
            try:
                sig = _sigpac_get(f"{SIGPAC_BASE}/recintos/{prov_cod}/{mun_cod}/0/0/{poligono}/{parcela}", timeout=4)
                feats = sig.get('features', [])
                if feats:
                    props = feats[0].get('properties', {})
                    dn = props.get('dn_surface')
                    if dn: sup_ha = round(float(dn) / 10000, 4)
                    uso_sigpac = props.get('uso_sigpac', '')
            except Exception:
                pass

        cur.execute('''INSERT INTO parcelas
            (user_id, nombre_finca, poligono, parcela_num, recinto,
             provincia_cod, provincia_nombre, municipio_cod, municipio_nombre,
             superficie_ha, uso_sigpac, activa)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,1)''',
            (uid, nombre, poligono, parcela, recinto or '1',
             prov_cod, prov_nom, mun_cod, mun_nom, sup_ha, uso_sigpac))
        n_ok += 1

    conn.commit(); conn.close()
    return n_ok, len(errores), errores


@app.route('/api/import/excel', methods=['POST'])
@login_required
def route_import_excel():
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': 'No se recibió ningún archivo'}), 400
    f = request.files['file']
    if not f.filename.endswith('.xlsx'):
        return jsonify({'ok': False, 'error': 'El archivo debe ser .xlsx'}), 400
    try:
        from openpyxl import load_workbook
        wb = load_workbook(f, data_only=True)
    except ImportError:
        return jsonify({'ok': False, 'error': 'openpyxl no instalado'}), 500
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Error al leer el archivo: {e}'}), 400

    uid = get_uid()
    try:
        n_ok, n_err, errores = _importar_parcelas_wb(wb, uid)
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Error interno al importar: {e}'}), 500
    msg = f"{n_ok} parcelas importadas"
    if n_err: msg += f", {n_err} filas con error"
    return jsonify({'ok': True, 'total': n_ok, 'resumen': msg, 'errores': errores})

@app.route('/api/import/gsheet', methods=['POST'])
@login_required
def route_import_gsheet():
    import re, urllib.request, io
    data = request.get_json() or {}
    url = (data.get('url') or '').strip()
    if not url:
        return jsonify({'ok': False, 'error': 'URL vacía'}), 400
    m = re.search(r'/spreadsheets/d/([a-zA-Z0-9_-]+)', url)
    if not m:
        return jsonify({'ok': False, 'error': 'URL de Google Sheets no válida'}), 400
    sheet_id = m.group(1)
    export_url = f'https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx'
    try:
        req = urllib.request.Request(export_url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=15) as resp:
            content = resp.read()
    except Exception as e:
        return jsonify({'ok': False, 'error': f'No se pudo descargar la hoja. ¿Está compartida públicamente? ({e})'}), 400
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Error al leer el archivo: {e}'}), 400

    uid = get_uid()
    try:
        n_ok, n_err, errores = _importar_parcelas_wb(wb, uid)
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Error interno al importar: {e}'}), 500
    msg = f"{n_ok} parcelas importadas"
    if n_err: msg += f", {n_err} filas con error"
    return jsonify({'ok': True, 'total': n_ok, 'resumen': msg, 'errores': errores})

# ─────────────────────────────────────────────
# EXPORT
# ─────────────────────────────────────────────
@app.route('/api/export/excel')
@login_required
@requires_active_plan
def route_export_excel():
    from exports import export_excel
    uid = get_uid()
    campana = request.args.get('campana', '2025/2026')
    return export_excel(uid, campana)

@app.route('/api/export/pdf')
@login_required
@requires_active_plan
def route_export_pdf():
    from export_pdf import export_pdf
    uid = get_uid()
    campana = request.args.get('campana', '2025/2026')
    return export_pdf(uid, campana)

@app.route('/api/backup/export')
@login_required
@admin_required
def backup_export():
    db_path = os.path.join(os.path.dirname(__file__), 'cuaderno.db')
    return send_file(db_path, as_attachment=True, download_name='cuaderno_backup.db')

@app.route('/api/backup/import', methods=['POST'])
@login_required
@admin_required
def backup_import():
    if 'file' not in request.files:
        return jsonify({"error": "Sin archivo"}), 400
    f = request.files['file']

    # Verificar magic bytes de SQLite antes de tocar nada
    header = f.read(16)
    if not header.startswith(b'SQLite format 3'):
        return jsonify({"error": "El archivo no es una base de datos SQLite válida"}), 400
    f.seek(0)

    db_path = os.path.join(os.path.dirname(__file__), 'cuaderno.db')
    tmp_path = db_path + '.import_tmp'
    backup_path = db_path + '.bak'

    f.save(tmp_path)
    try:
        # Doble verificación: intentar abrir el fichero temporal como SQLite
        import sqlite3 as _sq
        _sq.connect(tmp_path).close()
    except Exception as e:
        os.remove(tmp_path)
        return jsonify({"error": f"Base de datos inválida: {e}"}), 400

    # Backup del archivo actual antes de sobreescribir
    import shutil
    shutil.copy2(db_path, backup_path)
    os.replace(tmp_path, db_path)
    return jsonify({"status": "ok", "backup_saved": backup_path})


# ─────────────────────────────────────────────
# METEOROLOGÍA — PROXY ALERTAS AEMET
# ─────────────────────────────────────────────
@app.route('/api/aemet/alertas')
@login_required
def aemet_alertas():
    import xml.etree.ElementTree as ET
    provincia = (request.args.get('provincia') or '').strip().lower()
    api_key = os.environ.get('AEMET_API_KEY', '')
    if not api_key:
        return jsonify({'ok': True, 'alertas': [], 'msg': 'AEMET_API_KEY no configurada'})
    try:
        r1 = req_lib.get(
            'https://opendata.aemet.es/opendata/api/avisos_cap/ultimoelaborado',
            headers={'api_key': api_key},
            timeout=10
        )
        if r1.status_code != 200:
            return jsonify({'ok': False, 'alertas': []})
        datos_url = r1.json().get('datos')
        if not datos_url:
            return jsonify({'ok': False, 'alertas': []})

        r2 = req_lib.get(datos_url, timeout=15)
        root = ET.fromstring(r2.text)

        CAP  = 'urn:oasis:names:tc:emergency:cap:1.2'
        ATOM = 'http://www.w3.org/2005/Atom'
        severity_map = {
            'Minor':    ('amarillo', '🟡'),
            'Moderate': ('naranja',  '🟠'),
            'Severe':   ('rojo',     '🔴'),
            'Extreme':  ('rojo',     '🔴'),
        }

        def parse_alert(alert_el):
            results = []
            for info in alert_el.findall(f'{{{CAP}}}info'):
                lang = info.findtext(f'{{{CAP}}}language') or ''
                if lang and 'es' not in lang.lower():
                    continue
                event    = info.findtext(f'{{{CAP}}}event')    or ''
                severity = info.findtext(f'{{{CAP}}}severity') or 'Minor'
                headline = info.findtext(f'{{{CAP}}}headline') or event
                expires  = info.findtext(f'{{{CAP}}}expires')  or ''
                nivel, icono = severity_map.get(severity, ('amarillo', '🟡'))
                for area in info.findall(f'{{{CAP}}}area'):
                    area_desc = area.findtext(f'{{{CAP}}}areaDesc') or ''
                    if provincia and provincia not in area_desc.lower():
                        continue
                    results.append({
                        'nivel': nivel, 'icon': icono,
                        'evento': event, 'area': area_desc,
                        'texto': headline, 'expira': expires,
                        'fuente': 'AEMET',
                    })
            return results

        alertas = []
        for entry in root.findall(f'{{{ATOM}}}entry'):
            for alert in entry.findall(f'{{{CAP}}}alert'):
                alertas.extend(parse_alert(alert))
        if not alertas:
            if root.tag == f'{{{CAP}}}alert':
                alertas.extend(parse_alert(root))
            else:
                for alert in root.findall(f'{{{CAP}}}alert'):
                    alertas.extend(parse_alert(alert))

        return jsonify({'ok': True, 'alertas': alertas})
    except Exception as e:
        return jsonify({'ok': False, 'alertas': [], 'error': str(e)})


# ─────────────────────────────────────────────
# STRIPE — PAGOS
# ─────────────────────────────────────────────

STRIPE_SECRET_KEY      = os.environ.get('STRIPE_SECRET_KEY', '')
STRIPE_WEBHOOK_SECRET  = os.environ.get('STRIPE_WEBHOOK_SECRET', '')
STRIPE_PRICES = {
    ('basic', 'monthly'): os.environ.get('STRIPE_PRICE_BASIC_MONTHLY', ''),
    ('basic', 'yearly'):  os.environ.get('STRIPE_PRICE_BASIC_YEARLY', ''),
    ('pro',   'monthly'): os.environ.get('STRIPE_PRICE_PRO_MONTHLY', ''),
    ('pro',   'yearly'):  os.environ.get('STRIPE_PRICE_PRO_YEARLY', ''),
}

def _stripe():
    """Devuelve el módulo stripe inicializado, o None si no hay clave configurada."""
    if not STRIPE_SECRET_KEY:
        return None
    import stripe as _s
    _s.api_key = STRIPE_SECRET_KEY
    return _s


def _plan_from_price(stripe_price_id):
    """Identifica el plan ('basic'/'pro') a partir del Price ID de Stripe."""
    for (plan, _), pid in STRIPE_PRICES.items():
        if pid and pid == stripe_price_id:
            return plan
    return 'basic'


@app.route('/api/stripe/checkout', methods=['POST'])
@login_required
def stripe_checkout():
    """Crea una sesión de Stripe Checkout y devuelve la URL de pago."""
    s = _stripe()
    if not s:
        return jsonify({"error": "Stripe no configurado"}), 503

    data    = request.json or {}
    plan    = data.get('plan', 'basic')
    billing = data.get('billing', 'monthly')

    price_id = STRIPE_PRICES.get((plan, billing))
    if not price_id:
        return jsonify({"error": "Plan o intervalo no válido"}), 400

    base_url = request.host_url.rstrip('/')

    conn = get_db()
    u = one(conn, "SELECT stripe_customer_id FROM users WHERE id=?", (current_user.id,))
    conn.close()
    customer_id = u.get('stripe_customer_id') if u else None

    try:
        params = {
            "mode": "subscription",
            "line_items": [{"price": price_id, "quantity": 1}],
            "success_url": f"{base_url}/pago-completado?session_id={{CHECKOUT_SESSION_ID}}",
            "cancel_url":  f"{base_url}/#planes",
            "metadata": {
                "user_id": str(current_user.id),
                "plan":    plan,
            },
            "subscription_data": {
                "metadata": {"user_id": str(current_user.id), "plan": plan}
            },
        }
        if customer_id:
            params["customer"] = customer_id
        else:
            params["customer_email"] = current_user.email

        session_obj = s.checkout.Session.create(**params)
        return jsonify({"url": session_obj.url})
    except Exception as e:
        app.logger.error("Stripe checkout error: %s", e)
        return jsonify({"error": "Error al crear sesión de pago"}), 500


@app.route('/api/stripe/portal', methods=['POST'])
@login_required
def stripe_portal():
    """Crea una sesión del portal de cliente de Stripe (gestión de suscripción)."""
    s = _stripe()
    if not s:
        return jsonify({"error": "Stripe no configurado"}), 503

    conn = get_db()
    u = one(conn, "SELECT stripe_customer_id FROM users WHERE id=?", (current_user.id,))
    conn.close()

    customer_id = u.get('stripe_customer_id') if u else None
    if not customer_id:
        return jsonify({"error": "No tienes una suscripción activa en Stripe"}), 400

    try:
        portal = s.billing_portal.Session.create(
            customer=customer_id,
            return_url=request.host_url,
        )
        return jsonify({"url": portal.url})
    except Exception as e:
        app.logger.error("Stripe portal error: %s", e)
        return jsonify({"error": "Error al abrir el portal"}), 500


@app.route('/api/stripe/webhook', methods=['POST'])
def stripe_webhook():
    """Recibe eventos de Stripe y actualiza el plan del usuario en BD."""
    s = _stripe()
    if not s:
        return jsonify({"error": "Stripe no configurado"}), 503

    payload = request.get_data()
    sig     = request.headers.get('Stripe-Signature', '')

    try:
        event = s.Webhook.construct_event(payload, sig, STRIPE_WEBHOOK_SECRET)
    except Exception as e:
        app.logger.error("Stripe webhook signature error: %s", e)
        return jsonify({"error": "Invalid signature"}), 400

    ev_type = event['type']
    obj     = event['data']['object']

    conn = get_db()
    try:
        if ev_type in ('customer.subscription.created', 'customer.subscription.updated'):
            sub        = obj
            customer   = sub.get('customer')
            status     = sub.get('status')
            items      = sub.get('items', {}).get('data', [])
            price_id   = items[0]['price']['id'] if items else None
            plan       = _plan_from_price(price_id)
            period_end = sub.get('current_period_end')
            sub_end    = (datetime.datetime.utcfromtimestamp(period_end).strftime('%Y-%m-%d %H:%M:%S')
                          if period_end else None)
            sub_id     = sub.get('id')
            uid_meta   = sub.get('metadata', {}).get('user_id')

            if uid_meta:
                if status == 'active':
                    conn.execute(
                        "UPDATE users SET plan=?, subscription_ends_at=?, stripe_customer_id=?, stripe_subscription_id=? WHERE id=?",
                        (plan, sub_end, customer, sub_id, int(uid_meta))
                    )
                elif status in ('canceled', 'unpaid', 'past_due'):
                    conn.execute(
                        "UPDATE users SET plan='trial', subscription_ends_at=NULL WHERE id=?",
                        (int(uid_meta),)
                    )
                conn.commit()

        elif ev_type == 'customer.subscription.deleted':
            sub      = obj
            uid_meta = sub.get('metadata', {}).get('user_id')
            if uid_meta:
                conn.execute(
                    "UPDATE users SET plan='trial', subscription_ends_at=NULL, stripe_subscription_id=NULL WHERE id=?",
                    (int(uid_meta),)
                )
                conn.commit()

        elif ev_type == 'checkout.session.completed':
            session_obj = obj
            customer    = session_obj.get('customer')
            uid_meta    = session_obj.get('metadata', {}).get('user_id')
            plan_meta   = session_obj.get('metadata', {}).get('plan')
            sub_id      = session_obj.get('subscription')
            if uid_meta and customer:
                if plan_meta in ('basic', 'pro'):
                    import datetime as _dt
                    sub_end = (_dt.datetime.utcnow() + _dt.timedelta(days=365)).strftime('%Y-%m-%d %H:%M:%S')
                    conn.execute(
                        "UPDATE users SET plan=?, stripe_customer_id=?, stripe_subscription_id=?, subscription_ends_at=? WHERE id=?",
                        (plan_meta, customer, sub_id, sub_end, int(uid_meta))
                    )
                else:
                    conn.execute(
                        "UPDATE users SET stripe_customer_id=? WHERE id=?",
                        (customer, int(uid_meta))
                    )
                conn.commit()
    finally:
        conn.close()

    return jsonify({"status": "ok"})


# ─────────────────────────────────────────────
# BOOT
# ─────────────────────────────────────────────
if __name__ == '__main__':
    init_db()
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_ENV') != 'production'
    app.run(host='0.0.0.0', port=port, debug=debug)
