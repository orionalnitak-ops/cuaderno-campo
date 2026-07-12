"""
blueprints/imports_exports.py — /api/import/*, /api/export/*, /api/backup/*
"""
import io
import logging
import os

import requests as req_lib
from flask import Blueprint, jsonify, request, send_file
from flask_login import login_required
from helpers import get_uid, admin_required, requires_active_plan, get_active_explotacion_id
from blueprints.sigpac import _sigpac_get, SIGPAC_BASE
from extensions import limiter

bp = Blueprint('imports_exports', __name__)
logger = logging.getLogger(__name__)


def _importar_parcelas_wb(wb, uid):
    """Importa parcelas desde un workbook openpyxl. Devuelve (n_importadas, n_errores, detalle_errores)."""
    import unicodedata
    from db import get_db

    def _norm(s):
        return unicodedata.normalize('NFKD', str(s or '').lower().strip()).encode('ascii', 'ignore').decode()

    # Buscar hoja de parcelas (acepta cualquier nombre que contenga "parcela")
    ws = None
    for name in wb.sheetnames:
        if 'parcela' in _norm(name) or name == wb.sheetnames[0]:
            ws = wb[name]; break
    if not ws:
        return 0, 0, []

    # Leer cabeceras (fila 1) y mapear columnas por nombre
    headers = [_norm(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

    def _col(keywords):
        for kw in keywords:
            for i, h in enumerate(headers):
                if kw in h: return i
        return None

    ci_nombre    = _col(['nombre', 'finca', 'parcela nom'])
    ci_comunidad = _col(['comunidad', 'ccaa', 'autonoma', 'autonomia'])
    ci_provincia = _col(['provincia', 'prov'])
    ci_municipio = _col(['municipio', 'muni'])
    ci_poligono  = _col(['poligon', 'polig'])
    ci_parcela   = _col(['num parcela', 'parcela', 'numero', 'parcela num'])
    ci_recinto   = _col(['recinto'])

    faltan = [n for n, c in [('Nombre', ci_nombre), ('Provincia', ci_provincia), ('Municipio', ci_municipio), ('Polígono', ci_poligono), ('Parcela', ci_parcela)] if c is None]
    if faltan:
        return 0, 0, [f"Faltan columnas: {', '.join(faltan)}"]

    # Cache de códigos provincia/municipio para no consultar SIGPAC en cada fila
    conn = get_db()
    cur = conn.cursor()
    exp_id = get_active_explotacion_id(conn)
    _cache_prov = {}; _cache_mun = {}

    def _get_prov_cod(nombre_prov):
        k = _norm(nombre_prov)
        if k not in _cache_prov:
            data = _sigpac_get(f"{SIGPAC_BASE}/provincias", timeout=4)
            for f in data.get('features', []):
                p = f.get('properties', {})
                _cache_prov[_norm(p.get('nombre', ''))] = str(p.get('codigo', ''))
        return _cache_prov.get(k, '')

    def _get_mun_cod(prov_cod, nombre_mun):
        k = f"{prov_cod}:{_norm(nombre_mun)}"
        if k not in _cache_mun:
            data = _sigpac_get(f"{SIGPAC_BASE}/municipios/{prov_cod}", timeout=4)
            for f in data.get('features', []):
                p = f.get('properties', {})
                _cache_mun[f"{prov_cod}:{_norm(p.get('nombre', ''))}"] = str(p.get('codigo', ''))
        return _cache_mun.get(k, '')

    n_ok = 0; errores = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row): continue
        nombre    = str(row[ci_nombre]    or '').strip()
        comunidad = str(row[ci_comunidad] or '').strip() if ci_comunidad is not None else ''
        prov_nom  = str(row[ci_provincia] or '').strip()
        mun_nom   = str(row[ci_municipio] or '').strip()
        poligono  = str(row[ci_poligono]  or '').strip()
        parcela   = str(row[ci_parcela]   or '').strip()
        recinto   = str(row[ci_recinto]   or '').strip() if ci_recinto is not None else ''

        if not nombre or not poligono or not parcela or not prov_nom or not mun_nom:
            errores.append(f"Fila incompleta: '{nombre or '?'}'")
            continue

        prov_cod = _get_prov_cod(prov_nom)
        mun_cod  = _get_mun_cod(prov_cod, mun_nom) if prov_cod else ''

        # Lookup SIGPAC para superficie y uso
        sup_ha = None; uso_sigpac = ''
        if prov_cod and mun_cod:
            try:
                sig = _sigpac_get(f"{SIGPAC_BASE}/recintos/{prov_cod}/{mun_cod}/0/0/{poligono}/{parcela}", timeout=4)
                feats = sig.get('features', [])
                if feats:
                    props = feats[0].get('properties', {})
                    dn = props.get('dn_surface')
                    if dn: sup_ha = round(float(dn) / 10000, 4)
                    uso_sigpac = props.get('uso_sigpac', '')
            except Exception:
                pass

        cur.execute('''INSERT INTO parcelas
            (user_id, explotacion_id, nombre_finca, comunidad, poligono, parcela_num, recinto,
             provincia_cod, provincia_nombre, municipio_cod, municipio_nombre,
             superficie_ha, uso_sigpac, activa)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,1)''',
            (uid, exp_id, nombre, comunidad or None, poligono, parcela, recinto or None,
             prov_cod, prov_nom, mun_cod, mun_nom, sup_ha, uso_sigpac))
        n_ok += 1

    conn.commit(); conn.close()
    return n_ok, len(errores), errores


@bp.route('/api/import/excel', methods=['POST'])
@login_required
@limiter.limit("10 per minute")
def route_import_excel():
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': 'No se recibió ningún archivo'}), 400
    f = request.files['file']
    if not f.filename.endswith('.xlsx'):
        return jsonify({'ok': False, 'error': 'El archivo debe ser .xlsx'}), 400
    try:
        from openpyxl import load_workbook
        wb = load_workbook(f, data_only=True)
    except ImportError:
        return jsonify({'ok': False, 'error': 'openpyxl no instalado'}), 500
    except Exception as e:
        logger.error("import/excel load_workbook: %s", e)
        return jsonify({'ok': False, 'error': 'El archivo no es un Excel válido'}), 400

    uid = get_uid()
    try:
        n_ok, n_err, errores = _importar_parcelas_wb(wb, uid)
    except Exception as e:
        logger.error("import/excel _importar_parcelas_wb uid=%s: %s", uid, e)
        return jsonify({'ok': False, 'error': 'Error interno al importar. Revisa el formato del archivo.'}), 500
    msg = f"{n_ok} parcelas importadas"
    if n_err: msg += f", {n_err} filas con error"
    return jsonify({'ok': True, 'total': n_ok, 'resumen': msg, 'errores': errores})


@bp.route('/api/import/gsheet', methods=['POST'])
@login_required
@limiter.limit("10 per minute")
def route_import_gsheet():
    import re, urllib.request
    data = request.get_json() or {}
    url = (data.get('url') or '').strip()
    if not url:
        return jsonify({'ok': False, 'error': 'URL vacía'}), 400
    m = re.search(r'/spreadsheets/d/([a-zA-Z0-9_-]+)', url)
    if not m:
        return jsonify({'ok': False, 'error': 'URL de Google Sheets no válida'}), 400
    sheet_id = m.group(1)
    export_url = f'https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx'
    try:
        req = urllib.request.Request(export_url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=15) as resp:
            content = resp.read()
    except Exception as e:
        logger.error("import/gsheet download sheet_id=%s: %s", sheet_id, e)
        return jsonify({'ok': False, 'error': 'No se pudo descargar la hoja. ¿Está compartida públicamente?'}), 400
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        logger.error("import/gsheet load_workbook: %s", e)
        return jsonify({'ok': False, 'error': 'El archivo descargado no es un Excel válido'}), 400

    uid = get_uid()
    try:
        n_ok, n_err, errores = _importar_parcelas_wb(wb, uid)
    except Exception as e:
        logger.error("import/gsheet _importar_parcelas_wb uid=%s: %s", uid, e)
        return jsonify({'ok': False, 'error': 'Error interno al importar. Revisa el formato del archivo.'}), 500
    msg = f"{n_ok} parcelas importadas"
    if n_err: msg += f", {n_err} filas con error"
    return jsonify({'ok': True, 'total': n_ok, 'resumen': msg, 'errores': errores})


@bp.route('/api/export/excel')
@login_required
@requires_active_plan
def route_export_excel():
    from exports import export_excel
    uid = get_uid()
    campana = request.args.get('campana', '2025/2026')
    exp_id = get_active_explotacion_id()
    try:
        return export_excel(uid, campana, exp_id)
    except Exception as e:
        import traceback
        logger.error("export_excel uid=%s campana=%s error: %s\n%s", uid, campana, e, traceback.format_exc())
        return jsonify({"ok": False, "error": str(e), "type": type(e).__name__}), 500


@bp.route('/api/export/pdf')
@login_required
@requires_active_plan
def route_export_pdf():
    from export_pdf import export_pdf
    uid = get_uid()
    campana = request.args.get('campana', '2025/2026')
    exp_id = get_active_explotacion_id()
    return export_pdf(uid, campana, exp_id)


@bp.route('/api/backup/export')
@login_required
@admin_required
def backup_export():
    db_path = os.path.join(os.path.dirname(__file__), '..', 'cuaderno.db')
    db_path = os.path.abspath(db_path)
    return send_file(db_path, as_attachment=True, download_name='cuaderno_backup.db')


@bp.route('/api/backup/import', methods=['POST'])
@login_required
@admin_required
def backup_import():
    if 'file' not in request.files:
        return jsonify({"error": "Sin archivo"}), 400
    f = request.files['file']

    # Verificar magic bytes de SQLite antes de tocar nada
    header = f.read(16)
    if not header.startswith(b'SQLite format 3'):
        return jsonify({"error": "El archivo no es una base de datos SQLite válida"}), 400
    f.seek(0)

    db_path = os.path.join(os.path.dirname(__file__), '..', 'cuaderno.db')
    db_path = os.path.abspath(db_path)
    tmp_path = db_path + '.import_tmp'
    backup_path = db_path + '.bak'

    f.save(tmp_path)
    try:
        # Doble verificación: intentar abrir el fichero temporal como SQLite
        import sqlite3 as _sq
        _sq.connect(tmp_path).close()
    except Exception as e:
        os.remove(tmp_path)
        return jsonify({"error": f"Base de datos inválida: {e}"}), 400

    # Backup del archivo actual antes de sobreescribir
    import shutil
    shutil.copy2(db_path, backup_path)
    os.replace(tmp_path, db_path)
    return jsonify({"status": "ok", "backup_saved": backup_path})
